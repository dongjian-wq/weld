import os
import sys
import time
import cv2
import torch
import numpy as np
import torchvision.transforms as transforms
from torch import nn
from PIL import Image, ImageOps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import matplotlib.pyplot as plt
from datetime import datetime
from ultralytics import YOLO
import argparse
import re  # 用于汉字检测的正则表达式

# ============== 配置参数 ==============
# 边框裁剪配置 - 支持四个方向分别设置
BORDER_CROP_CONFIG = {
    'top_ratio': 0.04,      # 顶部裁剪比例：通常较少，保留表头信息
    'bottom_ratio': 0.04,   # 底部裁剪比例：稍多，去除底部边线
    'left_ratio': 0.07, # 左侧裁剪比例：适中，去除左边线
    'right_ratio': 0.06   # 右侧裁剪比例：适中，去除右边线
}

# 汉字检测配置（专门针对手写汉字优化，减少误检）
CHINESE_DETECTION_CONFIG = {
    # 基础参数
    'enable_chinese_detection': True,    # 是否启用汉字检测
    'detection_level': 'cell',          # 检测级别：'table'=表格级, 'cell'=单元格级
    'exclude_rows': [2],                 # 排除的行号（从1开始计数），默认排除第2行
    
    # 单元格级检测参数（平衡版本：确保汉字能检出，数字通过惩罚控制）
    'confidence_threshold': 0.65,       # 检测置信度阈值（0.65=65%，平衡汉字检出和数字过滤）
    'min_foreground_density': 0.08,     # 最小前景密度（进一步降低以包含稀疏汉字）
    'max_foreground_density': 0.85,     # 最大前景密度（提高上限）
    'min_area_threshold': 40,           # 最小连通域面积（进一步降低以包含小汉字）
    'enable_digit_filter': True,        # 是否启用数字过滤（强烈推荐）
    'enable_debug_info': False,         # 是否显示详细检测信息（调试用）
    'center_focus_enabled': True,       # 专注中心区域，避免边缘表格线干扰（重要优化）
    
    # 表格级检测参数（保持原有逻辑）
    'min_char_components': 1,            # 最少有效字符数量（至少1个高质量汉字）
    'min_foreground_ratio': 0.08,       # 最小前景密度（提高要求）
    'max_foreground_ratio': 0.25,       # 最大前景密度（降低上限，避免数字表格）
    'min_edge_density': 0.05,           # 最小边缘密度（提高要求）
    'min_local_complexity': 15,         # 最小局部复杂度（提高要求）
    'min_distance_std': 30,             # 字符分布最小标准差（提高要求）
    'min_indicators': 3,                # 最少满足指标数（提高到3，减少误检）
    
    # 兼容旧配置的参数（保持向后兼容）
    'min_contour_area': 300,            # 提高最小轮廓面积
    'min_char_size': 20,                # 提高最小字符尺寸
    'aspect_ratio_min': 0.4,            # 更严格的宽高比
    'aspect_ratio_max': 2.5,            # 更严格的宽高比
}

# 其他配置参数可以在此添加
# IMAGE_CONFIG = {
#     'target_size': (28, 28),
#     'padding': 4,
#     'confidence_threshold': 0.7
# }

# 新的MNIST反向汉字检测算法说明：
# 策略：利用已训练的MNIST模型，反向判断汉字
# 1. 先用MNIST模型预测每个单元格
# 2. 如果模型置信度很低（< 0.4），说明不是数字/字母
# 3. 再进行简单的内容检查，有内容就认为是汉字
#
# 优点：
# - 简单可靠，避免复杂的特征工程
# - 利用现有的训练好的模型
# - 置信度计算更直观（MNIST置信度越低，汉字置信度越高）
# - 减少误检和漏检
#
# 调整方法：
# 在detect_chinese_by_mnist_reverse函数中调整：
# - mnist_threshold: MNIST置信度阈值（默认0.4）
# - content_ratio阈值: 前景内容比例要求（默认0.02-0.8）
# - simple_chinese_check中的参数: 汉字特征检查阈值
# =====================================

# 配置matplotlib支持中文显示
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'SimSun', 'Arial Unicode MS',
                                   'DejaVu Sans']  # 优先使用的字体系列
plt.rcParams['axes.unicode_minus'] = False  # 解决保存图像时负号'-'显示为方块的问题
plt.rcParams['font.family'] = 'sans-serif'  # 使用sans-serif字体
# 设置matplotlib最大图形数量警告阈值
plt.rcParams['figure.max_open_warning'] = 10


# 设置中文字体的函数
def set_chinese_font():
    """
    尝试设置中文字体，针对不同操作系统选择不同字体
    """
    import platform
    system = platform.system()

    if system == 'Windows':
        fonts = ['SimHei', 'Microsoft YaHei', 'SimSun']
    elif system == 'Darwin':  # macOS
        fonts = ['PingFang SC', 'Heiti SC', 'STHeiti']
    else:  # Linux等其他系统
        fonts = ['WenQuanYi Zen Hei', 'WenQuanYi Micro Hei', 'Droid Sans Fallback']

    # 尝试设置字体
    for font in fonts:
        try:
            plt.rcParams['font.sans-serif'] = [font] + plt.rcParams['font.sans-serif']
            # 测试字体是否可用
            fig = plt.figure(figsize=(1, 1))
            plt.text(0.5, 0.5, '测试', fontsize=12, ha='center', va='center')
            plt.close(fig)
            # print(f"成功设置中文字体: {font}")
            return True
        except Exception as e:
            continue

    print("警告: 未能找到合适的中文字体，图像中的中文可能不会正确显示")
    return False


# 在程序开始时调用此函数
set_chinese_font()


def create_timestamped_folders():
    """
    创建带时间戳的文件夹路径
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime("%Y-%m-%d_%H.%M.%S")

    # 创建各个处理阶段的输出文件夹
    data_dir = os.path.join(base_dir, "data")
    cell_dir = os.path.join(data_dir, "cell", timestamp)
    processed_cells_dir = os.path.join(data_dir, "processed_cells", timestamp)
    output_dir = os.path.join(base_dir, "处理结果", timestamp)

    # 确保各个文件夹存在
    for dir_path in [data_dir, os.path.dirname(cell_dir), os.path.dirname(processed_cells_dir),
                     os.path.dirname(output_dir), cell_dir, processed_cells_dir, output_dir]:
        os.makedirs(dir_path, exist_ok=True)

    return {
        "timestamp": timestamp,
        "cell_dir": cell_dir,
        "processed_cells_dir": processed_cells_dir,
        "output_dir": output_dir
    }


def detect_and_crop_largest_table_cell(image_path, output_folder, model_path, min_cell_size=10, conf_threshold=0.2):
    """
    使用YOLO模型检测表格区域并裁剪所有检测到的区域
    """
    print(f"使用YOLO模型处理图像: {image_path}")
    original = cv2.imread(image_path)
    if original is None:
        raise FileNotFoundError(f"无法读取图像: {image_path}")

    # 加载YOLO模型
    try:
        model = YOLO(model_path)
        # print(f"成功加载YOLO模型: {model_path}")
    except Exception as e:
        print(f"加载YOLO模型失败: {e}")
        return None

    # 使用YOLO进行检测
    results = model.predict(source=image_path, imgsz=1000, conf=conf_threshold)
    
    # 获取原图文件名
    filename = os.path.splitext(os.path.basename(image_path))[0]
    
    saved_regions = []
    
    for i, result in enumerate(results):
        if result.boxes is not None and len(result.boxes) > 0:
            # 裁剪每个目标
            for j, box in enumerate(result.boxes.xyxy.cpu().numpy()):
                x1, y1, x2, y2 = box.astype(int)
                
                # 边界扩展，确保表格线完整
                img_h, img_w = original.shape[:2]
                expand_pixels = 10  # 扩展像素数
                
                # 扩展右边界和下边界，确保表格线完整
                x2 = min(img_w, x2 + expand_pixels)
                y2 = min(img_h, y2 + expand_pixels)
                
                w = x2 - x1
                h = y2 - y1
                
                # 跳过小于最小尺寸的检测区域
                if w < min_cell_size or h < min_cell_size:
                    continue
                
                cropped = original[y1:y2, x1:x2]
                
                # 生成唯一文件名
                save_path = os.path.join(output_folder, f"{filename}_obj{j}.png")
                cv2.imwrite(save_path, cropped)
                
                # print(f"Saved: {save_path}")
                saved_regions.append(cropped)

    if saved_regions:
        # 保存第一个区域作为largest_cell.png（保持兼容性）
        cv2.imwrite(os.path.join(output_folder, "largest_cell.png"), saved_regions[0])
        # print(f"YOLO检测到 {len(saved_regions)} 个区域，返回第一个区域")
        return saved_regions[0]
    else:
        print("YOLO未检测到任何有效区域")
        return None


def extract_cells_by_intersections(image, horizontal_lines, vertical_lines, output_folder):
    """
    基于水平线和垂直线的交点精确分割单元格
    :param image: 原始图像
    :param horizontal_lines: 水平线图像
    :param vertical_lines: 垂直线图像
    :param output_folder: 输出文件夹
    :return: (单元格列表, 坐标列表)
    """
    # 检测水平线的y坐标
    h_lines_y = []
    contours, _ = cv2.findContours(horizontal_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        center_y = y + h // 2  # 使用水平线的中心y坐标
        h_lines_y.append(center_y)
    
    # 去重并排序
    h_lines_y = sorted(list(set(h_lines_y)))
    
    # 检测垂直线的x坐标
    v_lines_x = []
    contours, _ = cv2.findContours(vertical_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        center_x = x + w // 2  # 使用垂直线的中心x坐标
        v_lines_x.append(center_x)
    
    # 去重并排序
    v_lines_x = sorted(list(set(v_lines_x)))
    
    # 根据交点分割单元格
    cells = []
    coords = []
    cell_index = 1
    
    for i in range(len(h_lines_y) - 1):
        for j in range(len(v_lines_x) - 1):
            y1, y2 = h_lines_y[i], h_lines_y[i + 1]
            x1, x2 = v_lines_x[j], v_lines_x[j + 1]
            
            # 添加小的边距避免包含表格线
            margin = 3
            y1_margin = y1 + margin
            y2_margin = y2 - margin
            x1_margin = x1 + margin
            x2_margin = x2 - margin
            
            # 确保坐标在图像范围内
            h, w = image.shape[:2]
            y1_margin = max(0, min(y1_margin, h))
            y2_margin = max(y1_margin, min(y2_margin, h))
            x1_margin = max(0, min(x1_margin, w))
            x2_margin = max(x1_margin, min(x2_margin, w))
            
            # 提取单元格
            if y2_margin > y1_margin and x2_margin > x1_margin:
                cell = image[y1_margin:y2_margin, x1_margin:x2_margin]
                
                if cell.size > 0:
                    cells.append(cell)
                    coords.append((x1_margin, y1_margin, x2_margin - x1_margin, y2_margin - y1_margin))
                    
                    # 保存单元格
                    out_path = os.path.join(output_folder, f"cell_{cell_index}.png")
                    cv2.imwrite(out_path, cell)
                    
                    cell_index += 1

    # 保存表格线交点分析图
    intersection_debug = create_intersection_debug_image(image, h_lines_y, v_lines_x, coords, output_folder)
    
    return cells, coords


def create_intersection_debug_image(image, h_lines_y, v_lines_x, coords, output_folder):
    """
    创建表格线交点分析调试图像
    """
    debug_img = image.copy()
    if len(debug_img.shape) == 2:
        debug_img = cv2.cvtColor(debug_img, cv2.COLOR_GRAY2BGR)
    
    # 绘制水平线
    for y in h_lines_y:
        cv2.line(debug_img, (0, y), (debug_img.shape[1], y), (0, 255, 0), 2)
        cv2.putText(debug_img, f"H{y}", (5, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1)
    
    # 绘制垂直线
    for x in v_lines_x:
        cv2.line(debug_img, (x, 0), (x, debug_img.shape[0]), (255, 0, 0), 2)
        cv2.putText(debug_img, f"V{x}", (x+5, 15), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 0, 0), 1)
    
    # 绘制单元格边界和编号
    for i, (x, y, w, h) in enumerate(coords):
        cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 0, 255), 2)
        cv2.putText(debug_img, str(i + 1), (x + 5, y + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 0, 255), 2)
    
    # 保存调试图像
    debug_path = os.path.join(output_folder, "step3_intersection_analysis.png")
    cv2.imwrite(debug_path, debug_img)
    
    return debug_img


def create_grid_visualization(image, coords, output_folder):
    """
    创建分割结果的网格可视化图像
    """
    vis_img = image.copy()
    if len(vis_img.shape) == 2:
        vis_img = cv2.cvtColor(vis_img, cv2.COLOR_GRAY2BGR)
    
    # 绘制所有单元格边界
    for i, (x, y, w, h) in enumerate(coords):
        # 绘制单元格边界
        cv2.rectangle(vis_img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        
        # 添加单元格编号
        text = str(i + 1)
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.6
        thickness = 2
        
        # 计算文字大小和位置
        text_size = cv2.getTextSize(text, font, font_scale, thickness)[0]
        text_x = x + (w - text_size[0]) // 2
        text_y = y + (h + text_size[1]) // 2
        
        # 添加白色背景
        cv2.rectangle(vis_img, (text_x - 2, text_y - text_size[1] - 2), 
                     (text_x + text_size[0] + 2, text_y + 2), (255, 255, 255), -1)
        
        # 添加文字
        cv2.putText(vis_img, text, (text_x, text_y), font, font_scale, (255, 0, 0), thickness)
    
    # 保存可视化图像
    vis_path = os.path.join(output_folder, "step3_grid_visualization.png")
    cv2.imwrite(vis_path, vis_img)
    
    return vis_img


def split_into_small_cells(cropped_cell, output_folder):
    """
    将大的表格区域分割为小单元格
    """
    # 步骤一：图像预处理（灰度化和二值化）
    gray = cv2.cvtColor(cropped_cell, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, -2)
    
    # 保存步骤一的结果
    cv2.imwrite(os.path.join(output_folder, "step1_gray.png"), gray)
    cv2.imwrite(os.path.join(output_folder, "step1_thresh.png"), thresh)

    # 步骤二：形态学处理（提取表格线）
    horizontal = thresh.copy()
    vertical = thresh.copy()
    rows, cols = horizontal.shape
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (cols // 20, 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, rows // 20))
    horizontal = cv2.dilate(cv2.erode(horizontal, horizontal_kernel), horizontal_kernel)
    vertical = cv2.dilate(cv2.erode(vertical, vertical_kernel), vertical_kernel)

    # 保存原始垂直线
    cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_original.png"), vertical)
    
    # 连接断开的垂直线段（修复被手写数字1断开的表格线）
    def connect_vertical_lines(vertical_img, x_tolerance=15, gap_tolerance=80):
        """
        智能连接属于同一列的垂直线段，修复被手写数字断开的表格线
        优化算法：增强同列识别、动态间距判断、多轮连接
        """
        # 找到垂直线的轮廓
        contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # 创建带标记的彩色图像用于调试
        debug_img = cv2.cvtColor(vertical_img, cv2.COLOR_GRAY2BGR)
        
        # 创建新的垂直线图像
        connected_vertical = np.zeros_like(vertical_img)
        
        # 获取所有垂直线段的信息
        line_segments = []
        for i, cnt in enumerate(contours):
            x, y, w, h = cv2.boundingRect(cnt)
            if h > 3:  # 降低最小高度要求，包含更多短线段
                line_segments.append({
                    'id': i,
                    'x': x, 'y': y, 'w': w, 'h': h,
                    'center_x': x + w/2,
                    'left': x,
                    'right': x + w,
                    'top': y,
                    'bottom': y + h,
                    'used': False
                })
        
        # 在调试图像上标记原始线段
        for i, seg in enumerate(line_segments):
            x, y, w, h = seg['x'], seg['y'], seg['w'], seg['h']
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 255, 0), 1)
            cv2.putText(debug_img, f"{i}:({x},{y})", (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 0, 0), 1)
            cv2.putText(debug_img, f"{w}x{h}", (x, y + h + 12), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 0, 255), 1)
        
        # 智能分组：使用多种策略识别同一列的线段
        def is_same_column(seg1, seg2, tolerance):
            """判断两个线段是否属于同一列"""
            # 策略1：中心点x坐标接近
            center_diff = abs(seg1['center_x'] - seg2['center_x'])
            if center_diff <= tolerance:
                return True
            
            # 策略2：边界重叠（适用于宽度不同但位置相近的线段）
            overlap_left = max(seg1['left'], seg2['left'])
            overlap_right = min(seg1['right'], seg2['right'])
            overlap_width = max(0, overlap_right - overlap_left)
            
            # 如果有重叠且重叠宽度占较小线段的50%以上
            min_width = min(seg1['w'], seg2['w'])
            if overlap_width >= min_width * 0.5:
                return True
            
            # 策略3：较宽线段包含较窄线段
            if (seg1['left'] <= seg2['left'] <= seg2['right'] <= seg1['right']) or \
               (seg2['left'] <= seg1['left'] <= seg1['right'] <= seg2['right']):
                return True
            
            return False
        
        # 按x坐标分组，找到属于同一列的线段
        columns = []
        for seg in line_segments:
            if seg['used']:
                continue
            
            # 创建新的列
            column = [seg]
            seg['used'] = True
            
            # 查找同一列的其他线段（多轮查找，确保不遗漏）
            found_new = True
            while found_new:
                found_new = False
                for other_seg in line_segments:
                    if other_seg['used']:
                        continue
                    
                    # 检查是否与该列中的任何线段属于同一列
                    for col_seg in column:
                        if is_same_column(col_seg, other_seg, x_tolerance):
                            column.append(other_seg)
                            other_seg['used'] = True
                            found_new = True
                            break
                    
                    if found_new:
                        break
            
            if column:
                # 按y坐标排序
                column.sort(key=lambda item: item['top'])
                columns.append(column)
        
        # 连接每一列的线段
        connected_count = 0
        for col_idx, column in enumerate(columns):
            if len(column) <= 1:
                # 单个线段，直接绘制
                seg = column[0]
                cv2.rectangle(connected_vertical, (seg['x'], seg['y']), 
                            (seg['x'] + seg['w'], seg['y'] + seg['h']), 255, -1)
                continue
            
            # 计算列的整体参数
            min_x = min(seg['x'] for seg in column)
            max_x = max(seg['x'] + seg['w'] for seg in column)
            min_y = min(seg['top'] for seg in column)
            max_y = max(seg['bottom'] for seg in column)
            
            # 计算平均宽度和总高度
            avg_width = int(sum(seg['w'] for seg in column) / len(column))
            total_segment_height = sum(seg['h'] for seg in column)
            column_span = max_y - min_y
            
            # 动态判断是否连接（考虑多个因素）
            need_connect = False
            
            # 因素1：线段间距离
            max_gap = 0
            for i in range(len(column) - 1):
                gap = column[i+1]['top'] - column[i]['bottom']
                max_gap = max(max_gap, gap)
            
            # 因素2：线段密度（线段总高度与列跨度的比例）
            density_ratio = total_segment_height / column_span if column_span > 0 else 0
            
            # 智能连接判断
            if len(column) >= 3:  # 3个或更多线段，更倾向于连接
                if max_gap <= gap_tolerance * 1.5 or density_ratio >= 0.3:
                    need_connect = True
            elif len(column) == 2:  # 2个线段，标准判断
                if max_gap <= gap_tolerance or density_ratio >= 0.5:
                    need_connect = True
            
            # 因素3：如果线段数量多且总体看起来像一条列线，即使有较大间距也连接
            if len(column) >= 4 and density_ratio >= 0.25:
                need_connect = True
            
            if need_connect:
                # 绘制连接的完整垂直线，使用平均宽度避免过粗
                line_x = min_x
                line_width = max(int(avg_width), 2)  # 只使用平均宽度，确保最小宽度为2
                # 如果水平跨度太大，使用中心位置减少粗细
                if (max_x - min_x) > avg_width * 3:
                    # 水平跨度太大时，使用中心位置绘制
                    center_x = (min_x + max_x) / 2
                    line_x = int(center_x - line_width / 2)
                
                cv2.rectangle(connected_vertical, (line_x, min_y), 
                            (line_x + line_width, max_y), 255, -1)
                connected_count += 1
            else:
                # 间距太大或密度太低，分别绘制
                for seg in column:
                    cv2.rectangle(connected_vertical, (seg['x'], seg['y']), 
                                (seg['x'] + seg['w'], seg['y'] + seg['h']), 255, -1)
        
        # 保存带标记的调试图像
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_labeled.png"), debug_img)
        
        # 保存连接后的中间结果
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_connected.png"), connected_vertical)
        
        # 添加长度过滤：去除过短的垂直线（稍微放宽过滤条件）
        filtered_vertical = filter_short_vertical_lines(connected_vertical, 
                                                       output_folder,
                                                       min_length_ratio=0.70, 
                                                       min_absolute_length=150)
        
        return filtered_vertical
    
    def filter_short_vertical_lines(vertical_img, output_folder, min_length_ratio=0.15, min_absolute_length=30):
        """
        过滤掉过短的垂直线，保留主要的表格列线
        :param vertical_img: 垂直线图像
        :param output_folder: 输出文件夹路径
        :param min_length_ratio: 最小长度比例（相对于图像高度）
        :param min_absolute_length: 最小绝对长度（像素）
        """
        img_height, img_width = vertical_img.shape
        min_length_by_ratio = int(img_height * min_length_ratio)
        min_length = max(min_length_by_ratio, min_absolute_length)
        
        # 找到所有垂直线轮廓
        contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # 创建过滤后的图像
        filtered_img = np.zeros_like(vertical_img)
        
        kept_count = 0
        filtered_count = 0
        
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            
            if h >= min_length:
                # 保留长度足够的垂直线
                cv2.drawContours(filtered_img, [cnt], -1, 255, -1)
                kept_count += 1
            else:
                # 过滤掉太短的垂直线
                filtered_count += 1
        
        # 保存过滤结果用于调试
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_filtered.png"), filtered_img)
        
        # 标准化垂直线长度：与最左侧第一条垂直线长度一致
        normalized_img = normalize_vertical_lines_length(filtered_img, output_folder)
        
        # 距离过滤：去除多余的垂直线
        distance_filtered_img = filter_redundant_vertical_lines(normalized_img, output_folder)
        
        # 添加缺失的垂直线：在间距过大的地方补充垂直线
        enhanced_img = add_missing_vertical_lines(distance_filtered_img, output_folder)
        
        return enhanced_img
    
    def normalize_vertical_lines_length(vertical_img, output_folder):
        """
        将所有垂直线的长度标准化为与最左侧第一条垂直线长度一致
        :param vertical_img: 垂直线图像
        :param output_folder: 输出文件夹路径
        """
        # 找到所有垂直线轮廓
        contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours) == 0:
            print("没有找到垂直线，跳过长度标准化")
            return vertical_img
        
        # 获取所有垂直线的信息并按x坐标排序
        lines_info = []
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            lines_info.append({
                'contour': cnt,
                'x': x, 'y': y, 'w': w, 'h': h,
                'center_x': x + w/2
            })
        
        # 按x坐标排序，找到最左侧的垂直线
        lines_info.sort(key=lambda item: item['center_x'])
        reference_line = lines_info[0]
        
        # 获取参考线的长度和y坐标范围
        ref_top = reference_line['y']
        ref_bottom = reference_line['y'] + reference_line['h']
        ref_height = reference_line['h']
        
        # 创建标准化后的图像
        normalized_img = np.zeros_like(vertical_img)
        
        # 重新绘制所有垂直线，使用统一的长度
        for i, line_info in enumerate(lines_info):
            x, y, w, h = line_info['x'], line_info['y'], line_info['w'], line_info['h']
            
            # 使用参考线的y坐标范围，但保持各自的x坐标和宽度
            cv2.rectangle(normalized_img, (x, ref_top), (x + w, ref_bottom), 255, -1)
        
        # 保存标准化结果
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_normalized.png"), normalized_img)
        
        return normalized_img
    
    def filter_redundant_vertical_lines(vertical_img, output_folder):
        """
        过滤多余的垂直线：当与左右相邻线的距离之和接近理想间距时过滤
        :param vertical_img: 垂直线图像
        :param output_folder: 输出文件夹路径
        """
        # 找到所有垂直线轮廓
        contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours) <= 2:
            return vertical_img
        
        # 获取所有垂直线的中心x坐标
        lines_info = []
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            lines_info.append({
                'contour': cnt,
                'x': x, 'y': y, 'w': w, 'h': h,
                'center_x': x + w/2
            })
        
        # 按x坐标排序
        lines_info.sort(key=lambda item: item['center_x'])
        
        # 计算第一条线到最后一条线的总距离，除以16得到理想间距
        total_distance = lines_info[-1]['center_x'] - lines_info[0]['center_x']
        ideal_spacing = total_distance / 16
        tolerance = ideal_spacing * 0.2  # 20%的容差
        
        # 检查每条线（除了第一条和最后一条）
        lines_to_keep = []
        filtered_count = 0
        
        for i, line_info in enumerate(lines_info):
            # 始终保留第一条和最后一条线
            if i == 0 or i == len(lines_info) - 1:
                lines_to_keep.append(line_info)
                continue
            
            # 计算与左右相邻线的距离之和
            left_distance = line_info['center_x'] - lines_info[i-1]['center_x']
            right_distance = lines_info[i+1]['center_x'] - line_info['center_x']
            total_adjacent_distance = left_distance + right_distance
            
            # 如果距离之和接近理想间距，则过滤该线
            distance_diff = abs(total_adjacent_distance - ideal_spacing)
            
            if distance_diff <= tolerance:
                # 过滤掉这条线
                filtered_count += 1
            else:
                # 保留这条线
                lines_to_keep.append(line_info)
        
        # 创建过滤后的图像
        filtered_img = np.zeros_like(vertical_img)
        
        # 绘制保留的垂直线
        for line_info in lines_to_keep:
            cv2.drawContours(filtered_img, [line_info['contour']], -1, 255, -1)
        
        # 保存过滤结果
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_redundancy_filtered.png"), filtered_img)
        
        return filtered_img
    
    def add_missing_vertical_lines(vertical_img, output_folder):
        """
        检查相邻垂直线间距，在间距过大的地方添加缺失的垂直线
        :param vertical_img: 垂直线图像
        :param output_folder: 输出文件夹路径
        """
        # 找到所有垂直线轮廓
        contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours) < 2:
            return vertical_img
        
        # 获取所有垂直线的信息
        lines_info = []
        for cnt in contours:
            x, y, w, h = cv2.boundingRect(cnt)
            lines_info.append({
                'contour': cnt,
                'x': x, 'y': y, 'w': w, 'h': h,
                'center_x': x + w/2
            })
        
        # 按x坐标排序
        lines_info.sort(key=lambda item: item['center_x'])
        
        # 计算相邻垂直线之间的距离
        distances = []
        for i in range(len(lines_info) - 1):
            distance = lines_info[i+1]['center_x'] - lines_info[i]['center_x']
            distances.append(distance)
        
        if not distances:
            return vertical_img
        
        # 找出距离相近的组，选择最常见的距离组
        def group_similar_distances(distances, tolerance=0.15):
            """将相近的距离分组，返回最大组的平均距离"""
            if not distances:
                return 0
            
            # 对距离进行排序
            sorted_distances = sorted(distances)
            groups = []
            current_group = [sorted_distances[0]]
            
            for i in range(1, len(sorted_distances)):
                # 如果当前距离与组内平均距离的差异小于容差
                group_mean = np.mean(current_group)
                if abs(sorted_distances[i] - group_mean) / group_mean <= tolerance:
                    current_group.append(sorted_distances[i])
                else:
                    # 开始新组
                    groups.append(current_group)
                    current_group = [sorted_distances[i]]
            
            # 添加最后一组
            if current_group:
                groups.append(current_group)
            
            # 找出最大的组（包含最多距离的组）
            largest_group = max(groups, key=len)
            return np.mean(largest_group), largest_group
        
        # 获取标准距离（最常见的距离组的平均值）
        standard_distance, standard_group = group_similar_distances(distances)
        
        # print(f"标准距离组包含 {len(standard_group)} 个距离，平均值: {standard_distance:.2f}")
        # print(f"标准距离组: {[f'{d:.2f}' for d in standard_group]}")
        
        # 定义异常距离的阈值（标准距离的1.6倍）
        threshold = standard_distance * 1.6
        
        # 找出垂直线中最细的宽度
        min_width = min(line['w'] for line in lines_info)
        # print(f"使用最细线条宽度: {min_width} 像素")
        
        # 找出需要添加垂直线的位置
        enhanced_img = vertical_img.copy()
        added_lines = []
        
        for i, distance in enumerate(distances):
            if distance > threshold:
                # 计算需要添加的垂直线数量
                num_segments = int(round(distance / standard_distance))
                if num_segments > 1:
                    # 在这个间距中添加垂直线
                    left_line = lines_info[i]
                    right_line = lines_info[i + 1]
                    
                    # 使用左右两条线的平均参数作为新线的参数（除了宽度）
                    avg_y = (left_line['y'] + right_line['y']) // 2
                    avg_h = (left_line['h'] + right_line['h']) // 2
                    # 使用最细的线条宽度
                    new_w = min_width
                    
                    # print(f"在位置 {i} 添加 {num_segments-1} 条垂直线，间距从 {distance:.2f} 分割为 {num_segments} 段")
                    
                    # 计算新垂直线的x坐标
                    for j in range(1, num_segments):
                        new_x = left_line['center_x'] + (right_line['center_x'] - left_line['center_x']) * j / num_segments
                        new_x = int(new_x - new_w / 2)  # 转换为左上角x坐标
                        
                        # 绘制新的垂直线
                        cv2.rectangle(enhanced_img, (new_x, avg_y), (new_x + new_w, avg_y + avg_h), 255, -1)
                        
                        added_lines.append({
                            'x': new_x,
                            'y': avg_y,
                            'w': new_w,
                            'h': avg_h,
                            'center_x': new_x + new_w/2
                        })
        
        # 创建调试图像显示添加的垂直线
        debug_img = cv2.cvtColor(enhanced_img, cv2.COLOR_GRAY2BGR)
        
        # 用不同颜色标记原始线条和新增线条
        for line_info in lines_info:
            x, y, w, h = line_info['x'], line_info['y'], line_info['w'], line_info['h']
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 255, 0), 2)  # 绿色：原始线条
            cv2.putText(debug_img, f"原{w}", (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 255, 0), 1)
        
        for line_info in added_lines:
            x, y, w, h = line_info['x'], line_info['y'], line_info['w'], line_info['h']
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 0, 255), 2)  # 红色：新增线条
            cv2.putText(debug_img, f"新{w}", (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 0, 255), 1)
        
        # 在调试图像上添加距离信息
        for i, distance in enumerate(distances):
            left_x = int(lines_info[i]['center_x'])
            right_x = int(lines_info[i + 1]['center_x'])
            mid_x = (left_x + right_x) // 2
            mid_y = debug_img.shape[0] // 2
            
            # 根据距离是否异常选择颜色
            is_small = distance < 35  # 使用35像素作为判断标准
            color = (0, 0, 255) if is_small else (255, 0, 0)  # 红色：小间距，蓝色：正常
            cv2.putText(debug_img, f"{distance:.1f}", (mid_x, mid_y), cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1)
        
        # 保存调试图像
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_debug_added.png"), debug_img)
        
        # 保存增强后的垂直线图像
        cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines_enhanced.png"), enhanced_img)
        
        # print(f"添加了 {len(added_lines)} 条垂直线")
        
        return enhanced_img
    
    # 过滤水平线：去除长度远小于第一条横线长度的线段
    def filter_short_horizontal_lines(horizontal_img, output_folder):
        """
        过滤掉过短的水平线段，基于最长水平线的长度
        :param horizontal_img: 水平线图像
        :param output_folder: 输出文件夹路径
        """
        # 找到所有水平线轮廓
        contours, _ = cv2.findContours(horizontal_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours) == 0:
            return horizontal_img
        
        # 收集所有水平线段的宽度信息
        horizontal_segments = []
        for i, cnt in enumerate(contours):
            x, y, w, h = cv2.boundingRect(cnt)
            if w > 3:  # 宽度大于3像素的线段
                horizontal_segments.append({
                    'contour': cnt,
                    'x': x, 'y': y, 'w': w, 'h': h,
                    'width': w
                })
        
        if len(horizontal_segments) == 0:
            return horizontal_img
        
        # 找到最长的水平线作为参考
        max_width = max(seg['width'] for seg in horizontal_segments)
        min_width_threshold = max_width * 0.3  # 最小长度为最长线的30%
        
        # 创建过滤后的图像
        filtered_horizontal = np.zeros_like(horizontal_img)
        
        kept_count = 0
        filtered_count = 0
        
        for seg in horizontal_segments:
            if seg['width'] >= min_width_threshold:
                # 保留长度足够的水平线
                cv2.drawContours(filtered_horizontal, [seg['contour']], -1, 255, -1)
                kept_count += 1
            else:
                # 过滤掉太短的水平线
                filtered_count += 1
        
        return filtered_horizontal
    
    # 修复中断和过短的水平线
    def repair_horizontal_lines(horizontal_img, output_folder, y_tolerance=5, max_gap=100):
        """
        修复中断和过短的水平线
        :param horizontal_img: 水平线图像
        :param output_folder: 输出文件夹路径
        :param y_tolerance: 同一行的y坐标容差
        :param max_gap: 最大连接间距
        """
        # 找到所有水平线轮廓
        contours, _ = cv2.findContours(horizontal_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        if len(contours) == 0:
            return horizontal_img
        
        # 收集所有水平线段信息，并计算厚度统计
        all_segments = []
        for i, cnt in enumerate(contours):
            x, y, w, h = cv2.boundingRect(cnt)
            if w > 5:  # 宽度大于5像素的线段
                all_segments.append({
                    'id': i,
                    'x': x, 'y': y, 'w': w, 'h': h,
                    'center_y': y + h/2,
                    'left': x,
                    'right': x + w
                })
        
        # 计算主要行线的厚度范围，过滤掉明显过细的线
        if len(all_segments) > 0:
            heights = [seg['h'] for seg in all_segments]
            # 使用中位数作为标准厚度
            heights.sort()
            median_height = heights[len(heights)//2]
            min_height_threshold = max(2, median_height * 0.5)  # 最小厚度为中位数的50%
            
            # 只保留厚度合理的线段
            line_segments = []
            for seg in all_segments:
                if seg['h'] >= min_height_threshold:
                    line_segments.append(seg)
        else:
            line_segments = all_segments
        
        # 按y坐标分组，找到同一行的线段
        rows = []
        used = [False] * len(line_segments)
        
        for i, seg in enumerate(line_segments):
            if used[i]:
                continue
            
            # 创建新的行，包含当前线段
            row = [seg]
            used[i] = True
            
            # 找到y坐标相近的其他线段
            for j, other_seg in enumerate(line_segments):
                if used[j] or i == j:
                    continue
                
                # 检查y坐标是否相近
                if abs(seg['center_y'] - other_seg['center_y']) <= y_tolerance:
                    row.append(other_seg)
                    used[j] = True
            
            # 按x坐标排序
            row.sort(key=lambda x: x['left'])
            rows.append(row)
        
        # 创建修复后的图像
        repaired_img = np.zeros_like(horizontal_img)
        
        # 找到最长的行作为参考长度
        max_row_width = 0
        leftmost_x = float('inf')
        rightmost_x = 0
        
        for row in rows:
            if len(row) > 0:
                row_left = min(seg['left'] for seg in row)
                row_right = max(seg['right'] for seg in row)
                row_width = row_right - row_left
                
                leftmost_x = min(leftmost_x, row_left)
                rightmost_x = max(rightmost_x, row_right)
                max_row_width = max(max_row_width, row_width)
        
        # 修复每一行 - 强制所有行线统一到最大长度
        repaired_count = 0
        for row_idx, row in enumerate(rows):
            if len(row) == 1:
                # 单个线段，强制延长到标准长度
                seg = row[0]
                
                # 所有单个线段都延长到标准长度（从左边界到右边界）
                # 使用全局的左右边界确保所有行线长度一致
                cv2.rectangle(repaired_img, (leftmost_x, seg['y']), 
                            (rightmost_x, seg['y'] + seg['h']), 255, -1)
                
                repaired_count += 1
            
            elif len(row) > 1:
                # 多个线段，检查是否需要连接
                total_gaps = 0
                for i in range(len(row) - 1):
                    gap = row[i+1]['left'] - row[i]['right']
                    total_gaps += gap
                
                avg_gap = total_gaps / (len(row) - 1) if len(row) > 1 else 0
                
                # 如果平均间距不太大，连接成完整线条
                if avg_gap <= max_gap:
                    min_y = min(seg['y'] for seg in row)
                    max_y = max(seg['y'] + seg['h'] for seg in row)
                    
                    # 连接后的行线也强制延长到统一长度（从左边界到右边界）
                    cv2.rectangle(repaired_img, (leftmost_x, min_y), 
                                (rightmost_x, max_y), 255, -1)
                    
                    repaired_count += 1
                else:
                    # 间距太大，但仍然统一长度绘制
                    min_y = min(seg['y'] for seg in row)
                    max_y = max(seg['y'] + seg['h'] for seg in row)
                    
                    # 即使间距大也强制统一长度
                    cv2.rectangle(repaired_img, (leftmost_x, min_y), 
                                (rightmost_x, max_y), 255, -1)
        
        # 保存修复结果
        cv2.imwrite(os.path.join(output_folder, "step2_horizontal_lines_repaired.png"), repaired_img)
        
        return repaired_img
    
    # 应用水平线修复和过滤
    horizontal_repaired = repair_horizontal_lines(horizontal, output_folder, y_tolerance=5, max_gap=100)
    horizontal_filtered = filter_short_horizontal_lines(horizontal_repaired, output_folder)
    
    # 应用垂直线连接和过滤
    vertical_processed = connect_vertical_lines(vertical, x_tolerance=8, gap_tolerance=80)
    
    # 使用处理后的水平线和垂直线
    table_lines = cv2.bitwise_or(horizontal_filtered, vertical_processed)
    
    # 保存步骤二的结果
    cv2.imwrite(os.path.join(output_folder, "step2_horizontal_lines.png"), horizontal_filtered)
    cv2.imwrite(os.path.join(output_folder, "step2_vertical_lines.png"), vertical_processed)
    cv2.imwrite(os.path.join(output_folder, "step2_table_lines.png"), table_lines)

    # 在step2_horizontal_lines.png生成后立即分析水平线间距（静默模式）
    spacing_analysis = analyze_horizontal_lines_spacing(horizontal_filtered, output_folder, verbose=False)

    # 分析垂直线弯曲度（只分析长度>=150像素的垂直线，静默模式）
    vertical_curvature_analysis = analyze_vertical_lines_curvature(vertical, output_folder, min_length_threshold=150, verbose=False)

    # 步骤三：基于表格线交点的精确分割
    small_cells, cell_coords = extract_cells_by_intersections(
        cropped_cell, horizontal_filtered, vertical_processed, output_folder
    )
    
    # 创建分割结果可视化图像
    grid_visualization = create_grid_visualization(cropped_cell, cell_coords, output_folder)

    # 创建单元格网格图像用于比对
    create_cells_grid(small_cells, output_folder)
    
    return list(small_cells)


def create_cells_grid(cells, output_folder, grid_cols=16, cell_size=(60, 60), padding=2):
    """
    将分割出的单元格拼接成网格图像，方便比对裁剪准确性
    :param cells: 单元格图像列表
    :param output_folder: 输出文件夹路径
    :param grid_cols: 每行显示的单元格数量（默认16列）
    :param cell_size: 每个单元格的显示大小
    :param padding: 单元格之间的间距
    """
    if not cells:
        print("没有单元格可以拼接")
        return
    
    import math
    
    num_cells = len(cells)
    grid_rows = math.ceil(num_cells / grid_cols)
    
    # 计算网格图像的总尺寸
    grid_width = grid_cols * cell_size[0] + (grid_cols + 1) * padding
    grid_height = grid_rows * cell_size[1] + (grid_rows + 1) * padding
    
    # 创建白色背景的网格图像
    grid_image = np.ones((grid_height, grid_width, 3), dtype=np.uint8) * 255
    
    for i, cell in enumerate(cells):
        row = i // grid_cols
        col = i % grid_cols
        
        # 计算当前单元格在网格中的位置
        x = col * (cell_size[0] + padding) + padding
        y = row * (cell_size[1] + padding) + padding
        
        # 调整单元格大小
        if len(cell.shape) == 2:  # 灰度图
            cell_resized = cv2.resize(cell, cell_size)
            cell_resized = cv2.cvtColor(cell_resized, cv2.COLOR_GRAY2BGR)
        else:  # 彩色图
            cell_resized = cv2.resize(cell, cell_size)
        
        # 将单元格放置到网格中
        grid_image[y:y+cell_size[1], x:x+cell_size[0]] = cell_resized
        
        # 在单元格上添加序号（从1开始）
        font = cv2.FONT_HERSHEY_SIMPLEX
        font_scale = 0.4
        color = (0, 0, 255)  # 红色
        thickness = 1
        text = str(i + 1)
        
        # 计算文字大小和位置
        text_size = cv2.getTextSize(text, font, font_scale, thickness)[0]
        text_x = x + 2
        text_y = y + text_size[1] + 2
        
        # 添加白色背景
        cv2.rectangle(grid_image, (text_x-1, text_y-text_size[1]-1), 
                     (text_x + text_size[0]+1, text_y+2), (255, 255, 255), -1)
        
        # 添加文字
        cv2.putText(grid_image, text, (text_x, text_y), font, font_scale, color, thickness)
    
    # 保存网格图像
    grid_path = os.path.join(output_folder, "cells_grid.png")
    cv2.imwrite(grid_path, grid_image)
    # print(f"单元格网格图像已保存: {grid_path}")
    
    # 同时保存一个带边框的版本，更清楚地显示每个单元格
    grid_with_border = grid_image.copy()
    for i in range(num_cells):
        row = i // grid_cols
        col = i % grid_cols
        
        x = col * (cell_size[0] + padding) + padding
        y = row * (cell_size[1] + padding) + padding
        
        # 绘制边框
        cv2.rectangle(grid_with_border, (x-1, y-1), (x+cell_size[0]+1, y+cell_size[1]+1), (0, 0, 0), 1)
    
    bordered_grid_path = os.path.join(output_folder, "cells_grid_with_borders.png")
    cv2.imwrite(bordered_grid_path, grid_with_border)
    # print(f"带边框的单元格网格图像已保存: {bordered_grid_path}")


def crop_border(image, border_ratio=0.1, top_ratio=None, bottom_ratio=None, left_ratio=None, right_ratio=None):
    """
    裁剪图片边缘，支持四个方向分别设置裁剪比例
    :param image: 输入图像
    :param border_ratio: 统一裁剪比例（当单独方向未设置时使用）
    :param top_ratio: 顶部裁剪比例
    :param bottom_ratio: 底部裁剪比例
    :param left_ratio: 左侧裁剪比例
    :param right_ratio: 右侧裁剪比例
    """
    h, w = image.shape[:2] if len(image.shape) > 2 else image.shape
    
    # 如果没有单独设置，则使用统一的border_ratio
    top_ratio = top_ratio if top_ratio is not None else border_ratio
    bottom_ratio = bottom_ratio if bottom_ratio is not None else border_ratio
    left_ratio = left_ratio if left_ratio is not None else border_ratio
    right_ratio = right_ratio if right_ratio is not None else border_ratio
    
    # 计算各方向的裁剪像素数
    crop_top = int(h * top_ratio)
    crop_bottom = int(h * bottom_ratio)
    crop_left = int(w * left_ratio)
    crop_right = int(w * right_ratio)
    
    # 确保裁剪后仍有有效区域
    crop_top = min(crop_top, h // 3)
    crop_bottom = min(crop_bottom, h // 3)
    crop_left = min(crop_left, w // 3)
    crop_right = min(crop_right, w // 3)
    
    # 确保裁剪范围合理
    if crop_top + crop_bottom >= h:
        crop_top = crop_bottom = h // 4
    if crop_left + crop_right >= w:
        crop_left = crop_right = w // 4

    if len(image.shape) > 2:
        cropped = image[crop_top:h - crop_bottom, crop_left:w - crop_right]
    else:
        cropped = image[crop_top:h - crop_bottom, crop_left:w - crop_right]

    return cropped


def remove_table_lines(image):
    """
    精确检测并去除图像中的表格线，保护数字内容
    :param image: 输入的灰度图像
    :return: 去除表格线后的图像
    """
    # 获取图像尺寸
    height, width = image.shape
    
    # 创建结果图像的副本
    result = image.copy()
    
    # 先进行二值化，便于检测线条
    _, binary = cv2.threshold(image, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    
    # 检测水平线条
    # 使用较大的水平核，只检测真正的水平线
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (max(40, width//10), 1))
    horizontal_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, horizontal_kernel, iterations=1)
    
    # 检测垂直线条  
    # 使用较大的垂直核，只检测真正的垂直线
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, max(40, height//10)))
    vertical_lines = cv2.morphologyEx(binary, cv2.MORPH_OPEN, vertical_kernel, iterations=1)
    
    # 合并水平和垂直线条
    detected_lines = cv2.bitwise_or(horizontal_lines, vertical_lines)
    
    # 过滤掉过短的线条（避免误删数字的一部分）
    # 找到线条的轮廓
    contours, _ = cv2.findContours(detected_lines, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    # 创建过滤后的线条图像
    filtered_lines = np.zeros_like(detected_lines)
    
    for contour in contours:
        # 获取轮廓的边界框
        x, y, w, h = cv2.boundingRect(contour)
        
        # 判断是否为真正的表格线
        # 水平线：宽度远大于高度，且长度占图像宽度的一定比例
        # 垂直线：高度远大于宽度，且长度占图像高度的一定比例
        is_horizontal_line = w > h * 8 and w > width * 0.3  # 宽度至少是高度的8倍，且占图像宽度30%以上
        is_vertical_line = h > w * 8 and h > height * 0.3   # 高度至少是宽度的8倍，且占图像高度30%以上
        
        if is_horizontal_line or is_vertical_line:
            # 这是一条真正的表格线，添加到过滤后的图像中
            cv2.drawContours(filtered_lines, [contour], -1, 255, -1)
    
    # 对检测到的线条进行轻微膨胀，确保完全覆盖
    if np.any(filtered_lines > 0):
        # 使用更小的膨胀核，减少对数字的影响
        dilate_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2, 2))
        dilated_lines = cv2.dilate(filtered_lines, dilate_kernel, iterations=1)
        
        # 将检测到的线条区域设置为白色（背景色）
        result[dilated_lines > 0] = 255
    
    return result


def remove_table_lines_small(image):
    """
    专门针对小尺寸(28x28)图像的表格线过滤
    用于多级缩放后的精确清理
    :param image: 输入的小尺寸图像(numpy数组)
    :return: 清理后的图像
    """
    if len(image.shape) == 3:
        # 如果是3通道，转为单通道
        if image.shape[2] == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image[:,:,0]
    else:
        gray = image.copy()
    
    height, width = gray.shape
    result = gray.copy()
    
    # 对于小尺寸图像，使用更精细的检测
    # 检测完整的水平线和垂直线
    
    # 检测水平线：扫描每一行
    for y in range(height):
        row = result[y, :]
        # 计算这一行的黑色像素密度
        if len(row) > 0:
            black_pixels = np.sum(row == 0)  # 黑色像素数量
            black_ratio = black_pixels / len(row)
            
            # 如果超过70%的像素是黑色，且这一行跨越了图像的大部分宽度
            if black_ratio > 0.7 and black_pixels > width * 0.6:
                # 检查是否真的是表格线（避误删数字的一部分）
                # 查看上下相邻行的情况
                is_table_line = True
                check_range = min(2, height//4)  # 检查范围
                
                for offset in range(1, check_range + 1):
                    # 检查上方
                    if y - offset >= 0:
                        upper_row = result[y - offset, :]
                        upper_black_ratio = np.sum(upper_row == 0) / len(upper_row)
                        if upper_black_ratio > 0.5:  # 如果上方也有很多黑色，可能是数字的一部分
                            is_table_line = False
                            break
                    
                    # 检查下方
                    if y + offset < height:
                        lower_row = result[y + offset, :]
                        lower_black_ratio = np.sum(lower_row == 0) / len(lower_row)
                        if lower_black_ratio > 0.5:  # 如果下方也有很多黑色，可能是数字的一部分
                            is_table_line = False
                            break
                
                # 如果确认是表格线，则清除
                if is_table_line:
                    result[y, :] = 255
    
    # 检测垂直线：扫描每一列
    for x in range(width):
        col = result[:, x]
        # 计算这一列的黑色像素密度
        if len(col) > 0:
            black_pixels = np.sum(col == 0)  # 黑色像素数量
            black_ratio = black_pixels / len(col)
            
            # 如果超过70%的像素是黑色，且这一列跨越了图像的大部分高度
            if black_ratio > 0.7 and black_pixels > height * 0.6:
                # 检查是否真的是表格线
                is_table_line = True
                check_range = min(2, width//4)  # 检查范围
                
                for offset in range(1, check_range + 1):
                    # 检查左侧
                    if x - offset >= 0:
                        left_col = result[:, x - offset]
                        left_black_ratio = np.sum(left_col == 0) / len(left_col)
                        if left_black_ratio > 0.5:
                            is_table_line = False
                            break
                    
                    # 检查右侧
                    if x + offset < width:
                        right_col = result[:, x + offset]
                        right_black_ratio = np.sum(right_col == 0) / len(right_col)
                        if right_black_ratio > 0.5:
                            is_table_line = False
                            break
                
                # 如果确认是表格线，则清除
                if is_table_line:
                    result[:, x] = 255
    
    return result



def preprocess_image(image, target_size=(28, 28), padding=4):
    """
    对图像进行预处理：裁剪边框、去除表格线、颜色反转、调整大小
    注意：暂时跳过二值化步骤，直接使用灰度图像观察效果
    """
    # 转为灰度图
    if len(image.shape) > 2:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image

    # 精确边框裁剪：使用配置参数分别设置四个方向的裁剪比例
    cropped = crop_border(gray, **BORDER_CROP_CONFIG)

    # 去除表格线
    no_lines = remove_table_lines(cropped)

    # 步骤4：二值化
    # 使用稍大的高斯模糊去噪
    blurred = cv2.GaussianBlur(no_lines, (5, 5), 0)
    
    # 使用自适应阈值进行二值化（调整参数让二值化更温和）
    binary = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, 5)
    
    # 步骤5：去除边缘残留表格线
    def remove_edge_table_lines_step(binary_img):
        if len(binary_img.shape) != 2:
            return binary_img
        
        height, width = binary_img.shape
        result = binary_img.copy()
        edge_thickness = max(3, min(height//10, width//10))  # 动态调整边缘厚度
        
        def analyze_horizontal_line_thickness(img, y_pos, max_thickness=2):
            """
            分析水平线在垂直方向上的厚度
            """
            if y_pos >= img.shape[0]:
                return False, 0
            
            # 检查当前行是否有黑色像素
            current_row = img[y_pos, :]
            if np.sum(current_row == 0) < width * 0.1:  # 至少10%是黑色
                return False, 0
            
            # 向上和向下扩展，计算连续的黑色行数
            thickness = 1  # 当前行
            
            # 向上检查
            for offset in range(1, max_thickness + 2):
                if y_pos - offset >= 0:
                    row = img[y_pos - offset, :]
                    if np.sum(row == 0) > width * 0.1:  # 仍有较多黑色
                        thickness += 1
                    else:
                        break
                else:
                    break
            
            # 向下检查
            for offset in range(1, max_thickness + 2):
                if y_pos + offset < img.shape[0]:
                    row = img[y_pos + offset, :]
                    if np.sum(row == 0) > width * 0.1:  # 仍有较多黑色
                        thickness += 1
                    else:
                        break
                else:
                    break
            
            # 判断是否为细线
            is_thin = thickness <= max_thickness
            return is_thin, thickness
        
        def analyze_vertical_line_thickness(img, x_pos, max_thickness=2):
            """
            分析垂直线在水平方向上的厚度
            """
            if x_pos >= img.shape[1]:
                return False, 0
            
            # 检查当前列是否有黑色像素
            current_col = img[:, x_pos]
            if np.sum(current_col == 0) < height * 0.1:  # 至少10%是黑色
                return False, 0
            
            # 向左和向右扩展，计算连续的黑色列数
            thickness = 1  # 当前列
            
            # 向左检查
            for offset in range(1, max_thickness + 2):
                if x_pos - offset >= 0:
                    col = img[:, x_pos - offset]
                    if np.sum(col == 0) > height * 0.1:  # 仍有较多黑色
                        thickness += 1
                    else:
                        break
                else:
                    break
            
            # 向右检查
            for offset in range(1, max_thickness + 2):
                if x_pos + offset < img.shape[1]:
                    col = img[:, x_pos + offset]
                    if np.sum(col == 0) > height * 0.1:  # 仍有较多黑色
                        thickness += 1
                    else:
                        break
                else:
                    break
            
            # 判断是否为细线
            is_thin = thickness <= max_thickness
            return is_thin, thickness
        
        # 1. 清除上边缘的水平细线
        for y in range(min(edge_thickness, height)):
            is_thin, thickness = analyze_horizontal_line_thickness(result, y, max_thickness=2)
            if is_thin:
                # 清除整个细线（包括相邻的行）
                for dy in range(-thickness//2, thickness//2 + 1):
                    if 0 <= y + dy < height:
                        result[y + dy, :] = 255
        
        # 2. 清除下边缘的水平细线
        for y in range(max(0, height - edge_thickness), height):
            is_thin, thickness = analyze_horizontal_line_thickness(result, y, max_thickness=2)
            if is_thin:
                for dy in range(-thickness//2, thickness//2 + 1):
                    if 0 <= y + dy < height:
                        result[y + dy, :] = 255
        
        # 3. 清除左边缘的垂直细线
        for x in range(min(edge_thickness, width)):
            is_thin, thickness = analyze_vertical_line_thickness(result, x, max_thickness=2)
            if is_thin:
                # 清除整个细线（包括相邻的列）
                for dx in range(-thickness//2, thickness//2 + 1):
                    if 0 <= x + dx < width:
                        result[:, x + dx] = 255
        
        # 4. 清除右边缘的垂直细线
        for x in range(max(0, width - edge_thickness), width):
            is_thin, thickness = analyze_vertical_line_thickness(result, x, max_thickness=2)
            if is_thin:
                for dx in range(-thickness//2, thickness//2 + 1):
                    if 0 <= x + dx < width:
                        result[:, x + dx] = 255
        
        return result
    
    # 应用边缘表格线清除
    denoised = remove_edge_table_lines_step(binary)
    
    # 轻微的形态学清理
    kernel_small = np.ones((2, 2), np.uint8)
    denoised = cv2.morphologyEx(denoised, cv2.MORPH_OPEN, kernel_small)
    
    # 使用去噪后的图像
    processed_img = denoised
    
    # 保存二值化处理后的图像（和其他步骤保存方式一致）
    # 这里暂时注释，实际保存在create_preprocessing_comparison_grids函数中统一处理
    # 如果需要单独保存二值化结果，可以在main函数中调用
    
    # 转为PIL图像
    pil_img = Image.fromarray(processed_img)

    # 步骤6：反转颜色（黑底白字）
    pil_img = ImageOps.invert(pil_img.convert('L'))

    # 步骤7：添加一些边距避免贴边
    pil_img = ImageOps.expand(pil_img, border=padding, fill=0)

    # 步骤8：专门清理单元格上方行线和左右垂直线段（在添加边距后、缩放前进行）
    img_array = np.array(pil_img)
    if img_array.size > 0:
        height, width = img_array.shape
        result = img_array.copy()
        
        # 1. 精准清理上方行线策略
        # 检测上方1/9区域的水平线条（精确定位）
        upper_region_height = max(1, height // 9)
        
        for y in range(upper_region_height):
            row = result[y, :]
            white_pixels = np.sum(row == 255)  # 纯白色像素（线条）
            
            # 更低的阈值，更敏感地检测行线
            if white_pixels > width * 0.25:  # 降低到25%
                # 简化判断逻辑：如果连续的白色像素很长，很可能是行线
                # 检查是否有长连续的白色段
                continuous_segments = []
                current_segment = 0
                
                for pixel in row:
                    if pixel == 255:  # 白色
                        current_segment += 1
                    else:  # 黑色
                        if current_segment > 0:
                            continuous_segments.append(current_segment)
                            current_segment = 0
                
                # 添加最后一个段
                if current_segment > 0:
                    continuous_segments.append(current_segment)
                
                # 如果有长连续段（占宽度20%以上），认为是行线
                max_segment = max(continuous_segments) if continuous_segments else 0
                if max_segment > width * 0.2:
                    # 这很可能是行线，清除整行
                    result[y, :] = 0  # 设为黑色背景
        
        # 2. 清理左右两边的垂直线段
        # 检测左侧1/9区域和右侧1/9区域的垂直线条
        left_region_width = max(1, width // 8)
        right_region_width = max(1, width // 9)
        
        # 清理左侧垂直线段
        for x in range(left_region_width):
            col = result[:, x]
            white_pixels = np.sum(col == 255)  # 纯白色像素（线条）
            
            if white_pixels > height * 0.25:  # 降低到25%
                # 检查是否有长连续的垂直白色段
                continuous_segments = []
                current_segment = 0
                
                for pixel in col:
                    if pixel == 255:  # 白色
                        current_segment += 1
                    else:  # 黑色
                        if current_segment > 0:
                            continuous_segments.append(current_segment)
                            current_segment = 0
                
                # 添加最后一个段
                if current_segment > 0:
                    continuous_segments.append(current_segment)
                
                # 如果有长连续段（占高度20%以上），认为是垂直线
                max_segment = max(continuous_segments) if continuous_segments else 0
                if max_segment > height * 0.2:
                    # 这很可能是垂直线，清除整列
                    result[:, x] = 0  # 设为黑色背景
        
        # 清理右侧垂直线段
        for x in range(width - right_region_width, width):
            col = result[:, x]
            white_pixels = np.sum(col == 255)  # 纯白色像素（线条）
            
            if white_pixels > height * 0.25:  # 降低到25%
                # 检查是否有长连续的垂直白色段
                continuous_segments = []
                current_segment = 0
                
                for pixel in col:
                    if pixel == 255:  # 白色
                        current_segment += 1
                    else:  # 黑色
                        if current_segment > 0:
                            continuous_segments.append(current_segment)
                            current_segment = 0
                
                # 添加最后一个段
                if current_segment > 0:
                    continuous_segments.append(current_segment)
                
                # 如果有长连续段（占高度20%以上），认为是垂直线
                max_segment = max(continuous_segments) if continuous_segments else 0
                if max_segment > height * 0.2:
                    # 这很可能是垂直线，清除整列
                    result[:, x] = 0  # 设为黑色背景
        
        result_img = Image.fromarray(result)
    else:
        # 如果没有处理结果，使用原图像
        result_img = pil_img

    # 步骤9：改进的缩放方法 - 保持数字特征
    pil_img = result_img
    
    # 计算等比例缩放尺寸，保持宽高比
    original_size = pil_img.size
    original_w, original_h = original_size
    target_w, target_h = 28, 28
    
    # 计算缩放比例，保持宽高比
    scale_w = target_w / original_w
    scale_h = target_h / original_h
    scale = min(scale_w, scale_h)  # 使用较小的缩放比例保持完整内容
    
    # 计算缩放后的实际尺寸
    new_w = int(original_w * scale)
    new_h = int(original_h * scale)
    
    # 改进的缩放方法
    def smart_resize(img, target_size, method='adaptive'):
        """
        智能缩放方法，针对数字特征优化
        :param img: PIL图像
        :param target_size: 目标尺寸 (width, height)
        :param method: 缩放方法 'adaptive', 'lanczos', 'bicubic', 'nearest'
        """
        if method == 'adaptive':
            # 自适应方法：根据图像内容选择最佳插值
            img_array = np.array(img)
            white_pixel_ratio = np.sum(img_array == 255) / img_array.size
            
            if white_pixel_ratio < 0.1:  # 内容很少，使用保守方法
                return img.resize(target_size, Image.Resampling.LANCZOS)
            elif white_pixel_ratio > 0.7:  # 内容很多，保持清晰
                return img.resize(target_size, Image.Resampling.NEAREST)
            else:  # 中等内容，平衡质量
                return img.resize(target_size, Image.Resampling.BICUBIC)
        
        elif method == 'lanczos':
            # Lanczos插值：高质量，适合保持细节
            return img.resize(target_size, Image.Resampling.LANCZOS)
        
        elif method == 'bicubic':
            # 双三次插值：平滑，适合数字
            return img.resize(target_size, Image.Resampling.BICUBIC)
        
        elif method == 'antialias':
            # 抗锯齿缩放：先稍微模糊再缩放
            if target_size[0] < img.size[0] or target_size[1] < img.size[1]:
                # 缩小时先轻微模糊
                img_array = np.array(img)
                blurred = cv2.GaussianBlur(img_array, (3, 3), 0.5)
                img = Image.fromarray(blurred)
            return img.resize(target_size, Image.Resampling.LANCZOS)
        
        else:  # 'nearest'
            return img.resize(target_size, Image.Resampling.NEAREST)
    
    # 使用改进的缩放方法
    scaled_pil = smart_resize(pil_img, (new_w, new_h), method='adaptive')
    current = np.array(scaled_pil)
    
    # 确保缩放后仍然是纯二值化
    current = (current > 127).astype(np.uint8) * 255

    # 步骤10：多级缩放后的精确表格线清理（针对二值化图像，保护数字"1"）
    # 在最终尺寸下进行更精确的线条检测和清理
    img_array = current
    if img_array.size > 0:  # 确保图像不为空
        # 对于二值化图像，检测纯白色线条（因为已经颜色反转）
        height, width = img_array.shape
        result = img_array.copy()
        
        # 首先检查整体图像是否可能包含数字"1"
        def is_potential_digit_one(image):
            """检查图像是否可能包含数字1"""
            total_white = np.sum(image == 255)
            total_pixels = image.size
            
            # 如果白色像素占比在合理范围内（数字1通常占比较少）
            if 0.05 <= total_white / total_pixels <= 0.4:
                # 检查是否有垂直集中的特征（数字1的特点）
                center_x = width // 2
                center_region_width = max(3, width // 4)
                left_bound = max(0, center_x - center_region_width // 2)
                right_bound = min(width, center_x + center_region_width // 2)
                
                center_white = np.sum(image[:, left_bound:right_bound] == 255)
                center_ratio = center_white / total_white if total_white > 0 else 0
                
                # 如果中心区域集中了大部分白色像素，可能是数字1
                return center_ratio > 0.6
            return False
        
        # 检查是否可能是数字1
        possible_digit_one = is_potential_digit_one(result)
        
        # 检测水平线条（纯白色线条，因为已经颜色反转）
        for y in range(height):
            row = result[y, :]
            white_pixels = np.sum(row == 255)  # 纯白色像素（字符）
            if white_pixels > width * 0.75:  # 提高到75%，更严格
                # 检查邻域避误删真正的字符
                is_line = True
                for offset in [1, 2]:
                    if y-offset >= 0:
                        upper_white = np.sum(result[y-offset, :] == 255)
                        if upper_white > width * 0.3:
                            is_line = False
                            break
                    if y+offset < height:
                        lower_white = np.sum(result[y+offset, :] == 255)
                        if lower_white > width * 0.3:
                            is_line = False
                            break
                if is_line:
                    result[y, :] = 0  # 设为黑色背景
        
        # 检测垂直线条（纯白色线条），增强数字1保护
        for x in range(width):
            col = result[:, x]
            white_pixels = np.sum(col == 255)  # 纯白色像素
            
            # 如果可能是数字1，提高阈值并增加额外检查
            if possible_digit_one:
                threshold = 0.85  # 更严格的阈值
                # 数字1保护：检查列是否在中心区域
                center_x = width // 2
                distance_from_center = abs(x - center_x)
                is_center_region = distance_from_center <= width // 3
                
                if is_center_region and white_pixels > height * 0.3:
                    # 在中心区域且有一定白色像素，很可能是数字1的一部分，跳过清理
                    continue
            else:
                threshold = 0.75  # 比原来稍微严格一点
            
            # 应用阈值检查
            if white_pixels > height * threshold:
                # 增强的邻域检查
                is_line = True
                
                # 检查更多的邻域偏移
                for offset in [1, 2, 3]:
                    if x-offset >= 0:
                        left_white = np.sum(result[:, x-offset] == 255)
                        if left_white > height * 0.25:  # 降低邻域阈值，更保守
                            is_line = False
                            break
                    if x+offset < width:
                        right_white = np.sum(result[:, x+offset] == 255)
                        if right_white > height * 0.25:  # 降低邻域阈值，更保守
                            is_line = False
                            break
                
                # 额外检查：如果可能是数字1，进行更严格的验证
                if possible_digit_one and is_line:
                    # 检查这一列是否包含图像的主要内容
                    col_importance = white_pixels / np.sum(result == 255) if np.sum(result == 255) > 0 else 0
                    if col_importance > 0.2:  # 如果这一列包含超过20%的图像内容，保留
                        is_line = False
                
                if is_line:
                    result[:, x] = 0  # 设为黑色背景
        
        current = result
    else:
        current = img_array

    # 步骤11：创建黑底图像并将缩放后的图片粘贴居中
    pil_img = Image.fromarray(current)
    new_img = Image.new('L', (28, 28), 0)
    upper_left = ((28 - pil_img.size[0]) // 2, (28 - pil_img.size[1]) // 2)
    new_img.paste(pil_img, upper_left)

    return new_img



def create_preprocessing_comparison_grids(cell_folder, output_folder):
    """
    为所有单元格创建预处理步骤对比网格图像
    显示每个预处理步骤对整个表格的影响
    """
    # 获取所有单元格文件（跳过cell_0）
    cell_files = []
    for file_name in os.listdir(cell_folder):
        if file_name.startswith('cell_') and file_name != 'cell_0.png' and file_name.endswith('.png'):
            cell_files.append(file_name)
    
    # 按数字排序
    def get_file_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')
    
    cell_files.sort(key=get_file_index)
    
    if not cell_files:
        print("没有找到有效的单元格文件")
        return
    
    # 加载所有单元格图像
    cell_images = []
    for file_name in cell_files:
        file_path = os.path.join(cell_folder, file_name)
        image = cv2.imread(file_path)
        if image is not None:
            cell_images.append(image)
    
    if not cell_images:
        print("没有成功加载任何单元格图像")
        return
    
    print(f"生成预处理步骤对比网格...")
    
    # 定义预处理步骤
    def apply_preprocessing_step(images, step_num):
        """对所有图像应用指定的预处理步骤（累积式处理）"""
        processed_images = []
        
        for image in images:
            # 累积式处理：每个步骤基于前一步骤的结果
            current = image.copy()
            
            # 步骤0：原始图像（什么都不做）
            if step_num >= 1:
                # 步骤1：灰度转换
                if len(current.shape) > 2:
                    current = cv2.cvtColor(current, cv2.COLOR_BGR2GRAY)
            
            if step_num >= 2:
                # 步骤2：边框裁剪
                current = crop_border(current, **BORDER_CROP_CONFIG)
            
            if step_num >= 3:
                # 步骤3：去除表格线
                current = remove_table_lines(current)
            
            if step_num >= 4:
                # 步骤4：二值化
                # 使用稍大的高斯模糊去噪
                blurred = cv2.GaussianBlur(current, (5, 5), 0)
                
                # 使用自适应阈值进行二值化（调整参数让二值化更温和）
                binary = cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, 5)
                
                current = binary
            
            if step_num >= 5:
                # 步骤5：去除边缘残留表格线
                def remove_edge_table_lines_step(binary_img):
                    if len(binary_img.shape) != 2:
                        return binary_img
                    
                    height, width = binary_img.shape
                    result = binary_img.copy()
                    edge_thickness = max(3, min(height//10, width//10))  # 动态调整边缘厚度
                    
                    def analyze_horizontal_line_thickness(img, y_pos, max_thickness=2):
                        """
                        分析水平线在垂直方向上的厚度
                        """
                        if y_pos >= img.shape[0]:
                            return False, 0
                        
                        # 检查当前行是否有黑色像素
                        current_row = img[y_pos, :]
                        if np.sum(current_row == 0) < width * 0.1:  # 至少10%是黑色
                            return False, 0
                        
                        # 向上和向下扩展，计算连续的黑色行数
                        thickness = 1  # 当前行
                        
                        # 向上检查
                        for offset in range(1, max_thickness + 2):
                            if y_pos - offset >= 0:
                                row = img[y_pos - offset, :]
                                if np.sum(row == 0) > width * 0.1:  # 仍有较多黑色
                                    thickness += 1
                                else:
                                    break
                            else:
                                break
                        
                        # 向下检查
                        for offset in range(1, max_thickness + 2):
                            if y_pos + offset < img.shape[0]:
                                row = img[y_pos + offset, :]
                                if np.sum(row == 0) > width * 0.1:  # 仍有较多黑色
                                    thickness += 1
                                else:
                                    break
                            else:
                                break
                        
                        # 判断是否为细线
                        is_thin = thickness <= max_thickness
                        return is_thin, thickness
                    
                    def analyze_vertical_line_thickness(img, x_pos, max_thickness=2):
                        """
                        分析垂直线在水平方向上的厚度
                        """
                        if x_pos >= img.shape[1]:
                            return False, 0
                        
                        # 检查当前列是否有黑色像素
                        current_col = img[:, x_pos]
                        if np.sum(current_col == 0) < height * 0.1:  # 至少10%是黑色
                            return False, 0
                        
                        # 向左和向右扩展，计算连续的黑色列数
                        thickness = 1  # 当前列
                        
                        # 向左检查
                        for offset in range(1, max_thickness + 2):
                            if x_pos - offset >= 0:
                                col = img[:, x_pos - offset]
                                if np.sum(col == 0) > height * 0.1:  # 仍有较多黑色
                                    thickness += 1
                                else:
                                    break
                            else:
                                break
                        
                        # 向右检查
                        for offset in range(1, max_thickness + 2):
                            if x_pos + offset < img.shape[1]:
                                col = img[:, x_pos + offset]
                                if np.sum(col == 0) > height * 0.1:  # 仍有较多黑色
                                    thickness += 1
                                else:
                                    break
                            else:
                                break
                        
                        # 判断是否为细线
                        is_thin = thickness <= max_thickness
                        return is_thin, thickness
                    
                    # 1. 清除上边缘的水平细线
                    for y in range(min(edge_thickness, height)):
                        is_thin, thickness = analyze_horizontal_line_thickness(result, y, max_thickness=2)
                        if is_thin:
                            # 清除整个细线（包括相邻的行）
                            for dy in range(-thickness//2, thickness//2 + 1):
                                if 0 <= y + dy < height:
                                    result[y + dy, :] = 255
                    
                    # 2. 清除下边缘的水平细线
                    for y in range(max(0, height - edge_thickness), height):
                        is_thin, thickness = analyze_horizontal_line_thickness(result, y, max_thickness=2)
                        if is_thin:
                            for dy in range(-thickness//2, thickness//2 + 1):
                                if 0 <= y + dy < height:
                                    result[y + dy, :] = 255
                    
                    # 3. 清除左边缘的垂直细线
                    for x in range(min(edge_thickness, width)):
                        is_thin, thickness = analyze_vertical_line_thickness(result, x, max_thickness=2)
                        if is_thin:
                            # 清除整个细线（包括相邻的列）
                            for dx in range(-thickness//2, thickness//2 + 1):
                                if 0 <= x + dx < width:
                                    result[:, x + dx] = 255
                    
                    # 4. 清除右边缘的垂直细线
                    for x in range(max(0, width - edge_thickness), width):
                        is_thin, thickness = analyze_vertical_line_thickness(result, x, max_thickness=2)
                        if is_thin:
                            for dx in range(-thickness//2, thickness//2 + 1):
                                if 0 <= x + dx < width:
                                    result[:, x + dx] = 255
                    
                    return result
                
                # 应用边缘表格线清除
                denoised = remove_edge_table_lines_step(current)
                
                # 轻微的形态学清理
                kernel_small = np.ones((2, 2), np.uint8)
                denoised = cv2.morphologyEx(denoised, cv2.MORPH_OPEN, kernel_small)
                
                current = denoised
            
            if step_num >= 6:
                # 步骤6：颜色反转
                current = cv2.bitwise_not(current)
            
            if step_num >= 7:
                # 步骤7：添加边距
                pil_img = Image.fromarray(current)
                padded_pil = ImageOps.expand(pil_img, border=4, fill=0)
                current = np.array(padded_pil)
            
            if step_num >= 8:
                # 步骤8：清理单元格上方行线和左右垂直线段
                if current.size > 0:
                    height, width = current.shape
                    result = current.copy()
                    
                    # 1. 精准清理上方行线策略
                    # 检测上方1/9区域的水平线条（精确定位）
                    upper_region_height = max(1, height // 9)
                    
                    for y in range(upper_region_height):
                        row = result[y, :]
                        white_pixels = np.sum(row == 255)  # 纯白色像素（线条）
                        
                        # 更低的阈值，更敏感地检测行线
                        if white_pixels > width * 0.25:  # 降低到25%
                            # 简化判断逻辑：如果连续的白色像素很长，很可能是行线
                            # 检查是否有长连续的白色段
                            continuous_segments = []
                            current_segment = 0
                            
                            for pixel in row:
                                if pixel == 255:  # 白色
                                    current_segment += 1
                                else:  # 黑色
                                    if current_segment > 0:
                                        continuous_segments.append(current_segment)
                                        current_segment = 0
                            
                            # 添加最后一个段
                            if current_segment > 0:
                                continuous_segments.append(current_segment)
                            
                            # 如果有长连续段（占宽度20%以上），认为是行线
                            max_segment = max(continuous_segments) if continuous_segments else 0
                            if max_segment > width * 0.2:
                                # 这很可能是行线，清除整行
                                result[y, :] = 0  # 设为黑色背景
                        
                    # 2. 清理左右两边的垂直线段
                    # 检测左侧1/9区域和右侧1/9区域的垂直线条
                    left_region_width = max(1, width // 9)
                    right_region_width = max(1, width // 9)
                    
                    # 清理左侧垂直线段
                    for x in range(left_region_width):
                        col = result[:, x]
                        white_pixels = np.sum(col == 255)  # 纯白色像素（线条）
                        
                        if white_pixels > height * 0.25:  # 降低到25%
                            # 检查是否有长连续的垂直白色段
                            continuous_segments = []
                            current_segment = 0
                            
                            for pixel in col:
                                if pixel == 255:  # 白色
                                    current_segment += 1
                                else:  # 黑色
                                    if current_segment > 0:
                                        continuous_segments.append(current_segment)
                                        current_segment = 0
                            
                            # 添加最后一个段
                            if current_segment > 0:
                                continuous_segments.append(current_segment)
                            
                            # 如果有长连续段（占高度20%以上），认为是垂直线
                            max_segment = max(continuous_segments) if continuous_segments else 0
                            if max_segment > height * 0.2:
                                # 这很可能是垂直线，清除整列
                                result[:, x] = 0  # 设为黑色背景
                        
                    # 清理右侧垂直线段
                    for x in range(width - right_region_width, width):
                        col = result[:, x]
                        white_pixels = np.sum(col == 255)  # 纯白色像素（线条）
                        
                        if white_pixels > height * 0.25:  # 降低到25%
                            # 检查是否有长连续的垂直白色段
                            continuous_segments = []
                            current_segment = 0
                            
                            for pixel in col:
                                if pixel == 255:  # 白色
                                    current_segment += 1
                                else:  # 黑色
                                    if current_segment > 0:
                                        continuous_segments.append(current_segment)
                                        current_segment = 0
                            
                            # 添加最后一个段
                            if current_segment > 0:
                                continuous_segments.append(current_segment)
                            
                            # 如果有长连续段（占高度20%以上），认为是垂直线
                            max_segment = max(continuous_segments) if continuous_segments else 0
                            if max_segment > height * 0.2:
                                # 这很可能是垂直线，清除整列
                                result[:, x] = 0  # 设为黑色背景
                        
                    current = result
            
            if step_num >= 9:
                # 步骤9：改进的缩放方法 - 保持数字特征
                pil_img = Image.fromarray(current)
                
                # 计算等比例缩放尺寸，保持宽高比
                original_size = pil_img.size
                original_w, original_h = original_size
                target_w, target_h = 28, 28
                
                # 计算缩放比例，保持宽高比
                scale_w = target_w / original_w
                scale_h = target_h / original_h
                scale = min(scale_w, scale_h)  # 使用较小的缩放比例保持完整内容
                
                # 计算缩放后的实际尺寸
                new_w = int(original_w * scale)
                new_h = int(original_h * scale)
                
                # 改进的缩放方法
                def smart_resize(img, target_size, method='adaptive'):
                    """
                    智能缩放方法，针对数字特征优化
                    :param img: PIL图像
                    :param target_size: 目标尺寸 (width, height)
                    :param method: 缩放方法 'adaptive', 'lanczos', 'bicubic', 'nearest'
                    """
                    if method == 'adaptive':
                        # 自适应方法：根据图像内容选择最佳插值
                        img_array = np.array(img)
                        white_pixel_ratio = np.sum(img_array == 255) / img_array.size
                        
                        if white_pixel_ratio < 0.1:  # 内容很少，使用保守方法
                            return img.resize(target_size, Image.Resampling.LANCZOS)
                        elif white_pixel_ratio > 0.7:  # 内容很多，保持清晰
                            return img.resize(target_size, Image.Resampling.NEAREST)
                        else:  # 中等内容，平衡质量
                            return img.resize(target_size, Image.Resampling.BICUBIC)
                    
                    elif method == 'lanczos':
                        # Lanczos插值：高质量，适合保持细节
                        return img.resize(target_size, Image.Resampling.LANCZOS)
                    
                    elif method == 'bicubic':
                        # 双三次插值：平滑，适合数字
                        return img.resize(target_size, Image.Resampling.BICUBIC)
                    
                    elif method == 'antialias':
                        # 抗锯齿缩放：先稍微模糊再缩放
                        if target_size[0] < img.size[0] or target_size[1] < img.size[1]:
                            # 缩小时先轻微模糊
                            img_array = np.array(img)
                            blurred = cv2.GaussianBlur(img_array, (3, 3), 0.5)
                            img = Image.fromarray(blurred)
                        return img.resize(target_size, Image.Resampling.LANCZOS)
                    
                    else:  # 'nearest'
                        return img.resize(target_size, Image.Resampling.NEAREST)
                
                # 使用改进的缩放方法
                scaled_pil = smart_resize(pil_img, (new_w, new_h), method='adaptive')
                current = np.array(scaled_pil)
                
                # 确保缩放后仍然是纯二值化
                current = (current > 127).astype(np.uint8) * 255
            
            if step_num >= 10:
                # 步骤10：小尺寸表格线精确清理（针对二值化图像，保护数字"1"）
                # 确保current变量已定义 - 如果步骤9没有执行，使用PIL图像转换
                try:
                    current.shape  # 测试current是否已定义且为numpy数组
                except (NameError, AttributeError):
                    current = np.array(pil_img)
                
                if current.size > 0:
                    height, width = current.shape
                    result = current.copy()
                    
                    # 首先检查整体图像是否可能包含数字"1"
                    def is_potential_digit_one(image):
                        """检查图像是否可能包含数字1"""
                        total_white = np.sum(image == 255)
                        total_pixels = image.size
                        
                        # 如果白色像素占比在合理范围内（数字1通常占比较少）
                        if 0.05 <= total_white / total_pixels <= 0.4:
                            # 检查是否有垂直集中的特征（数字1的特点）
                            center_x = width // 2
                            center_region_width = max(3, width // 4)
                            left_bound = max(0, center_x - center_region_width // 2)
                            right_bound = min(width, center_x + center_region_width // 2)
                            
                            center_white = np.sum(image[:, left_bound:right_bound] == 255)
                            center_ratio = center_white / total_white if total_white > 0 else 0
                            
                            # 如果中心区域集中了大部分白色像素，可能是数字1
                            return center_ratio > 0.6
                        return False
                    
                    # 检查是否可能是数字1
                    possible_digit_one = is_potential_digit_one(result)
                    
                    # 检测水平线条（纯白色线条，因为已经颜色反转）
                    for y in range(height):
                        row = result[y, :]
                        white_pixels = np.sum(row == 255)  # 纯白色像素（字符）
                        if white_pixels > width * 0.75:  # 提高到75%，更严格
                            # 检查邻域避误删真正的字符
                            is_line = True
                            for offset in [1, 2]:
                                if y-offset >= 0:
                                    upper_white = np.sum(result[y-offset, :] == 255)
                                    if upper_white > width * 0.3:
                                        is_line = False
                                        break
                                if y+offset < height:
                                    lower_white = np.sum(result[y+offset, :] == 255)
                                    if lower_white > width * 0.3:
                                        is_line = False
                                        break
                            if is_line:
                                result[y, :] = 0  # 设为黑色背景
                    
                    # 检测垂直线条（纯白色线条），增强数字1保护
                    for x in range(width):
                        col = result[:, x]
                        white_pixels = np.sum(col == 255)  # 纯白色像素
                        
                        # 如果可能是数字1，提高阈值并增加额外检查
                        if possible_digit_one:
                            threshold = 0.85  # 更严格的阈值
                            # 数字1保护：检查列是否在中心区域
                            center_x = width // 2
                            distance_from_center = abs(x - center_x)
                            is_center_region = distance_from_center <= width // 3
                            
                            if is_center_region and white_pixels > height * 0.3:
                                # 在中心区域且有一定白色像素，很可能是数字1的一部分，跳过清理
                                continue
                        else:
                            threshold = 0.75  # 比原来稍微严格一点
                        
                        # 应用阈值检查
                        if white_pixels > height * threshold:
                            # 增强的邻域检查
                            is_line = True
                            
                            # 检查更多的邻域偏移
                            for offset in [1, 2, 3]:
                                if x-offset >= 0:
                                    left_white = np.sum(result[:, x-offset] == 255)
                                    if left_white > height * 0.25:  # 降低邻域阈值，更保守
                                        is_line = False
                                        break
                                if x+offset < width:
                                    right_white = np.sum(result[:, x+offset] == 255)
                                    if right_white > height * 0.25:  # 降低邻域阈值，更保守
                                        is_line = False
                                        break
                            
                            # 额外检查：如果可能是数字1，进行更严格的验证
                            if possible_digit_one and is_line:
                                # 检查这一列是否包含图像的主要内容
                                col_importance = white_pixels / np.sum(result == 255) if np.sum(result == 255) > 0 else 0
                                if col_importance > 0.2:  # 如果这一列包含超过20%的图像内容，保留
                                    is_line = False
                            
                            if is_line:
                                result[:, x] = 0  # 设为黑色背景
                    
                    current = result
            
            if step_num >= 11:
                # 步骤11：居中放置到28x28
                # 确保current变量已定义
                try:
                    current.shape  # 测试current是否已定义且为numpy数组
                except (NameError, AttributeError):
                    current = np.array(pil_img)
                
                pil_img = Image.fromarray(current)
                new_img = Image.new('L', (28, 28), 0)
                upper_left = ((28 - pil_img.size[0]) // 2, (28 - pil_img.size[1]) // 2)
                new_img.paste(pil_img, upper_left)
                current = np.array(new_img)
            

            
            # 确保current变量存在，如果不存在则使用PIL图像
            try:
                current.shape  # 测试current是否已定义且为numpy数组
            except (NameError, AttributeError):
                current = np.array(pil_img)
            
            processed_images.append(current)
        
        return processed_images
    
    # 步骤名称
    step_names = [
        "0.原始图像",
        "1.灰度转换", 
        "2.边框裁剪",
        "3.去除表格线",
        "4.二值化",
        "5.去除边缘残留表格线",
        "6.颜色反转",
        "7.添加边距(4px)",
        "8.清理单元格上方行线和左右垂直线段",
        "9.改进的缩放方法(保持宽高比)",
        "10.小尺寸精确线条清理(二值化)",
        "11.居中放置(28x28)"
    ]
    
    # 为每个步骤生成网格图像
    for step_num, step_name in enumerate(step_names):
        # 对所有单元格应用当前步骤
        processed_images = apply_preprocessing_step(cell_images, step_num)
        
        # 创建网格图像
        grid_image = create_step_grid(processed_images, step_name, cell_size=60)
        
        # 保存网格图像
        grid_filename = f"step_{step_num:02d}_{step_name.replace('.', '_').replace('(', '_').replace(')', '_').replace('%', 'percent')}.png"
        grid_path = os.path.join(output_folder, grid_filename)
        cv2.imwrite(grid_path, grid_image)

def analyze_vertical_lines_curvature(vertical_img_path, output_folder, min_length_threshold=100, verbose=True):
    """
    分析step2_vertical_lines_original.png中每条垂直线的弯曲度状态
    只分析达到最小长度要求的垂直线
    """
    # 读取垂直线图像
    if isinstance(vertical_img_path, str):
        vertical_img = cv2.imread(vertical_img_path, cv2.IMREAD_GRAYSCALE)
        if vertical_img is None:
            print(f"无法读取图像: {vertical_img_path}")
            return []
    else:
        vertical_img = vertical_img_path
    
    # 找到所有垂直线轮廓
    contours, _ = cv2.findContours(vertical_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) == 0:
        if verbose:
            print("没有找到垂直线")
        return []
    
    # 先过滤出符合长度要求的垂直线
    long_lines = []
    for i, contour in enumerate(contours):
        x, y, w, h = cv2.boundingRect(contour)
        if h >= min_length_threshold:
            long_lines.append((i, contour, x, y, w, h))
    
    if verbose:
        print(f"\n=== 垂直线弯曲度分析 ===")
        print(f"共检测到 {len(contours)} 条垂直线")
        print(f"符合长度要求(>={min_length_threshold}px)的垂直线: {len(long_lines)} 条")
    
    if len(long_lines) == 0:
        if verbose:
            print("没有符合长度要求的垂直线")
        return []
    
    # 创建彩色调试图像
    debug_img = cv2.cvtColor(vertical_img, cv2.COLOR_GRAY2BGR)
    lines_analysis = []
    
    # 只分析较长的垂直线
    for idx, (original_idx, contour, x, y, w, h) in enumerate(long_lines):
        points = contour.reshape(-1, 2)
        points = points[points[:, 1].argsort()]
        
        if len(points) < 3:
            continue
        
        top_point = points[0]
        bottom_point = points[-1]
        line_length = np.sqrt((bottom_point[0] - top_point[0])**2 + (bottom_point[1] - top_point[1])**2)
        
        # 计算弯曲度
        deviations = []
        for point in points:
            a = bottom_point[1] - top_point[1]
            b = top_point[0] - bottom_point[0]
            c = bottom_point[0] * top_point[1] - top_point[0] * bottom_point[1]
            
            if a == 0 and b == 0:
                distance = 0
            else:
                distance = abs(a * point[0] + b * point[1] + c) / np.sqrt(a**2 + b**2)
            deviations.append(distance)
        
        max_deviation = max(deviations) if deviations else 0
        avg_deviation = np.mean(deviations) if deviations else 0
        straightness = max(0, 100 - max_deviation * 10)
        
        # 分类弯曲程度 - 放宽标准
        if max_deviation < 2.0:  # 从1.5放宽到2.0
            curvature_code = "0"  # 0=直线
            color = (0, 255, 0)
        elif max_deviation < 8.0:  # 从3.0放宽到8.0
            curvature_code = "1"  # 1=轻微弯曲
            color = (0, 255, 255)
        elif max_deviation < 15.0:  # 从6.0放宽到12.0
            curvature_code = "2"  # 2=中度弯曲
            color = (0, 165, 255)
        else:  # 只有偏差>13.0才算严重弯曲
            curvature_code = "3"  # 3=严重弯曲
            color = (0, 0, 255)
        
        line_info = {
            'index': idx + 1,
            'original_index': original_idx + 1,
            'x': x, 'y': y, 'w': w, 'h': h,
            'center_x': x + w/2,
            'length': line_length,
            'max_deviation': max_deviation,
            'avg_deviation': avg_deviation,
            'straightness': straightness,
            'curvature_code': curvature_code
        }
        
        lines_analysis.append(line_info)
        
        # 绘制分析结果（使用英文避免字体问题）
        cv2.drawContours(debug_img, [contour], -1, color, 3)
        cv2.line(debug_img, tuple(top_point.astype(int)), tuple(bottom_point.astype(int)), (255, 255, 255), 2)
        
        # 只显示数字编号和状态码
        label_x = max(5, x - 30)
        label_y = y + 20
        
        # 只使用数字，避免所有文字
        cv2.putText(debug_img, f"{idx+1}", (label_x, label_y), cv2.FONT_HERSHEY_SIMPLEX, 0.8, color, 2)
        cv2.putText(debug_img, f"{curvature_code}", (label_x, label_y + 25), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)
        cv2.putText(debug_img, f"{int(straightness)}", (label_x, label_y + 45), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 1)
    
    # 按x坐标排序
    lines_analysis.sort(key=lambda item: item['center_x'])
    
    # 输出分析结果
    if verbose:
        print(f"\n长垂直线弯曲度分析结果:")
        print("=" * 100)
        print(f"{'序号':<4} {'原序号':<6} {'X位置':<8} {'高度':<8} {'长度':<8} {'最大偏差':<10} {'平均偏差':<10} {'直线度':<8} {'状态':<12}")
        print("=" * 100)
        
        for line in lines_analysis:
            print(f"{line['index']:<4} {line['original_index']:<6} {line['center_x']:<8.1f} {line['h']:<8} "
                  f"{line['length']:<8.1f} {line['max_deviation']:<10.2f} {line['avg_deviation']:<10.2f} "
                  f"{line['straightness']:<8.0f}% {line['curvature_code']:<8}")
    
    # 保存分析结果图像
    analysis_path = os.path.join(output_folder, "long_vertical_lines_curvature_analysis.png")
    cv2.imwrite(analysis_path, debug_img)
    if verbose:
        print(f"\n长垂直线弯曲度分析图已保存: {analysis_path}")
    
    return lines_analysis
def create_step_grid(images, step_name, grid_cols=16, cell_size=60, padding=2):
    """
    将处理后的图像创建为网格布局
    """
    import math
    
    num_images = len(images)
    grid_rows = math.ceil(num_images / grid_cols)
    
    # 计算网格尺寸
    grid_width = grid_cols * cell_size + (grid_cols + 1) * padding
    title_height = 40
    grid_height = grid_rows * cell_size + (grid_rows + 1) * padding + title_height
    
    # 创建白色背景
    grid_image = np.ones((grid_height, grid_width, 3), dtype=np.uint8) * 255
    
    # 添加标题
    font = cv2.FONT_HERSHEY_SIMPLEX
    title_font_scale = 0.8
    title_thickness = 2

def analyze_horizontal_lines_spacing(horizontal_img, output_folder, verbose=True):
    """
    分析水平线之间的垂直距离
    """
    contours, _ = cv2.findContours(horizontal_img, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    if len(contours) < 2:
        if verbose:
            print("水平线数量不足，无法计算间距")
        return []
    
    lines_info = []
    for i, cnt in enumerate(contours):
        x, y, w, h = cv2.boundingRect(cnt)
        center_y = y + h/2
        lines_info.append({
            'index': i,
            'center_y': center_y,
            'x': x, 'y': y, 'w': w, 'h': h
        })
    
    lines_info.sort(key=lambda item: item['center_y'])
    
    distances = []
    if verbose:
        print(f"\n=== 水平线间距分析 ===")
        print(f"共检测到 {len(lines_info)} 条水平线")
    
    for i in range(len(lines_info) - 1):
        top_line = lines_info[i]
        bottom_line = lines_info[i + 1]
        distance = bottom_line['center_y'] - top_line['center_y']
        distances.append(distance)
        if verbose:
            print(f"第{i+1}条线 -> 第{i+2}条线: {distance:.2f} 像素")
    
    if distances:
        avg_distance = sum(distances) / len(distances)
        min_distance = min(distances)
        max_distance = max(distances)
        
        if verbose:
            print(f"\n距离统计:")
            print(f"平均距离: {avg_distance:.2f} 像素")
            print(f"最小距离: {min_distance:.2f} 像素")
            print(f"最大距离: {max_distance:.2f} 像素")
        
        # 检查是否有距离低于35的情况
        has_small_spacing = any(d < 35 for d in distances)
        if has_small_spacing and verbose:
            small_distances = [d for d in distances if d < 35]
            print(f"⚠️ 检测到 {len(small_distances)} 个小于35像素的间距: {[f'{d:.1f}' for d in small_distances]}")
        
        # 创建可视化图像
        debug_img = cv2.cvtColor(horizontal_img, cv2.COLOR_GRAY2BGR)
        
        for i, line_info in enumerate(lines_info):
            x, y, w, h = line_info['x'], line_info['y'], line_info['w'], line_info['h']
            center_y = int(line_info['center_y'])
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 255, 0), 2)
            cv2.putText(debug_img, f"H{i+1}", (x - 30, center_y + 5), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 1)
        
        for i, distance in enumerate(distances):
            top_y = int(lines_info[i]['center_y'])
            bottom_y = int(lines_info[i + 1]['center_y'])
            mid_y = (top_y + bottom_y) // 2
            mid_x = debug_img.shape[1] // 2
            
            # 根据距离选择颜色：小于40用红色，否则用蓝色
            color = (0, 0, 255) if distance < 40 else (255, 0, 0)
            cv2.line(debug_img, (mid_x, top_y), (mid_x, bottom_y), color, 2)
            cv2.putText(debug_img, f"{distance:.1f}", (mid_x + 10, mid_y), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.4, color, 1)
        
        analysis_path = os.path.join(output_folder, "step2_horizontal_lines_spacing_analysis.png")
        cv2.imwrite(analysis_path, debug_img)
        if verbose:
            print(f"间距分析图已保存: {analysis_path}")
        
        return {
            'distances': distances,
            'avg_distance': avg_distance,
            'min_distance': min_distance,
            'max_distance': max_distance,
            'lines_count': len(lines_info),
            'has_small_spacing': has_small_spacing
        }
    
    return []


def process_images_for_mnist(input_folder, output_folder):
    """
    处理所有单元格图像，使其适合MNIST模型识别
    """
    os.makedirs(output_folder, exist_ok=True)
    print(f"处理图像从 {input_folder} 到 {output_folder}")

    # 获取并排序输入文件
    file_list = []
    for file_name in os.listdir(input_folder):
        # 过滤掉非图片文件
        if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            continue

        # 过滤掉中间处理文件
        if file_name in ['cropped_cell.png', 'largest_cell.png']:
            # 仍然复制这些文件，但不进行处理
            input_path = os.path.join(input_folder, file_name)
            output_path = os.path.join(output_folder, file_name)
            try:
                image = cv2.imread(input_path)
                if image is not None:
                    cv2.imwrite(output_path, image)
            except Exception as e:
                print(f"复制 {file_name} 时出错: {e}")
            continue

        # 只处理cell_开头的文件，跳过cell_0.png
        if not file_name.startswith('cell_') or file_name == 'cell_0.png':
            continue

        file_list.append(file_name)

    # 对文件进行数值排序
    def get_file_index(filename):
        # 针对"cell_数字.png"格式的文件提取数字部分
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')  # 对于无法解析的文件名放在最后

    file_list.sort(key=get_file_index)

    processed_count = 0
    for file_name in file_list:
        input_path = os.path.join(input_folder, file_name)
        output_path = os.path.join(output_folder, file_name)

        try:
            # 读取图像
            image = cv2.imread(input_path)
            if image is None:
                print(f"无法读取图片: {input_path}")
                continue

            # 处理图像
            processed_img = preprocess_image(image)
            processed_img.save(output_path)
            processed_count += 1

        except Exception as e:
            print(f"处理 {file_name} 时出错: {e}")

    print(f"成功处理 {processed_count} 张图像")
    return processed_count


def load_model(model_path):
    """
    加载预训练的 MNIST 模型
    """
    class CNN(nn.Module):
        def __init__(self):
            super(CNN, self).__init__()
            # 第一个卷积块
            self.conv1 = nn.Sequential(
                nn.Conv2d(1, 32, kernel_size=3, padding=1),
                nn.BatchNorm2d(32),
                nn.ReLU(),
                nn.Conv2d(32, 64, kernel_size=3, padding=1),
                nn.BatchNorm2d(64),
                nn.ReLU(),
                nn.MaxPool2d(2)
            )
            
            # 第二个卷积块
            self.conv2 = nn.Sequential(
                nn.Conv2d(64, 128, kernel_size=3, padding=1),
                nn.BatchNorm2d(128),
                nn.ReLU(),
                nn.Conv2d(128, 128, kernel_size=3, padding=1),
                nn.BatchNorm2d(128),
                nn.ReLU(),
                nn.MaxPool2d(2)
            )
            
            # 全连接层
            self.fc = nn.Sequential(
                nn.Flatten(),
                nn.Linear(128 * 7 * 7, 256),
                nn.ReLU(),
                nn.Dropout(0.4),
                nn.Linear(256, 128),
                nn.ReLU(),
                nn.Dropout(0.3),
                nn.Linear(128, 14)  # 修改为14个类别（0-9数字 + N、X、G、S字母）
            )

        def forward(self, x):
            x = self.conv1(x)
            x = self.conv2(x)
            x = self.fc(x)
            return x

    model = CNN()
    model.load_state_dict(torch.load(model_path, map_location=torch.device('cuda')))
    model.eval()
    return model


def predict_digits(folder_path, model, transform):
    """
    使用模型预测文件夹中的手写数字和字母，筛选掉接近纯黑的图片
    """
    # 定义标签映射字典（更新为14个类别）
    label_map = {
        0: '0', 1: '1', 2: '2', 3: '3', 4: '4',
        5: '5', 6: '6', 7: '7', 8: '8', 9: '9',
        10: 'N', 11: 'X', 12: 'G', 13: 'S'  # 新的四个字母类别
    }
    
    results = {}
    skipped_files = {}  # 存储被跳过的文件及原因

    # 获取文件列表
    file_list = []
    for file_name in os.listdir(folder_path):
        # 过滤掉非图片文件
        if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            continue

        # 过滤掉中间处理文件
        if file_name in ['cropped_cell.png', 'largest_cell.png']:
            continue

        # 只处理cell_开头的文件，跳过cell_0.png
        if not file_name.startswith('cell_') or file_name == 'cell_0.png':
            continue

        file_list.append(file_name)

    # 对文件进行数值排序
    def get_file_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')

    # 按数值排序文件列表
    file_list.sort(key=get_file_index)

    # 处理排序后的文件
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)

        # 读取图片并转换为灰度图
        image = cv2.imread(file_path, cv2.IMREAD_GRAYSCALE)
        if image is None:
            print(f"无法读取图片: {file_path}")
            skipped_files[file_name] = "无法读取图片"
            continue

        # 检查图片是否接近纯黑
        if np.mean(image) < 5 or np.count_nonzero(image > 10) < image.shape[0] * image.shape[1] * 0.01:
            skipped_files[file_name] = "接近纯黑"
            continue

        # 确保图像是2维的灰度图（修复维度问题）
        if len(image.shape) == 3 and image.shape[2] == 1:
            image = image.squeeze(axis=2)
        
        # 转换为 PIL 格式并预处理
        image = Image.fromarray(image)
        image = transform(image).unsqueeze(0)  # 添加 batch 维度

        # 预测
        with torch.no_grad():
            output = model(image)
            # 计算概率分布和置信度
            probabilities = torch.softmax(output, dim=1)
            confidence, predicted_idx = torch.max(probabilities, dim=1)
            predicted_label = label_map[predicted_idx.item()]
            confidence_score = confidence.item()
            
            # 存储预测结果和置信度
            results[file_name] = {
                'label': predicted_label,
                'confidence': confidence_score
            }

    return results, skipped_files


def write_results_to_excel(results, excel_path, is_combined=False):
    """
    将预测结果写入 Excel 文件，按照固定的 10 行 16 列顺序排列
    除第一列外，剩下15列按5列一组分为3个类别，当某类别有置信度<0.65时，该类别整行标红
    """
    try:
        # 创建 Excel 工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "预测结果"

        # 记录需要标红的类别和行：{(row, category): True}
        red_categories = set()  # 存储需要标红的 (行号, 类别号)

        # 第一遍：收集所有需要标红的类别
        for file_name, result_info in results.items():
            try:
                # 获取文件索引
                if is_combined:
                    parts = file_name.split('_')
                    if len(parts) >= 3 and parts[-1].endswith('.png'):
                        index_str = parts[-1].split('.')[0]
                        if index_str.isdigit():
                            index = int(index_str)
                        else:
                            continue
                    else:
                        continue
                else:
                    if not file_name.startswith("cell_"):
                        continue
                    index = int(file_name.split("_")[1].split(".")[0])

                # 计算行列位置
                row = ((index - 1) // 16) + 1
                col = ((index - 1) % 16) + 1
                
                # 跳过第一列
                if col == 1:
                    continue
                    
                # 检查置信度
                if isinstance(result_info, dict) and 'confidence' in result_info:
                    confidence = result_info['confidence']
                    if confidence < 0.65:
                        # 确定类别：列2-6为类别1，列7-11为类别2，列12-16为类别3
                        if 2 <= col <= 6:  # 类别1 (列2-6)
                            category = 1
                        elif 7 <= col <= 11:  # 类别2 (列7-11)
                            category = 2
                        elif 12 <= col <= 16:  # 类别3 (列12-16)
                            category = 3
                        else:
                            continue
                        
                        # 标记该行的该类别需要标红
                        red_categories.add((row, category))

            except (IndexError, ValueError) as e:
                print(f"处理文件 {file_name} 时出错: {e}")
                continue

        # 移除调试打印
        # print(f"需要标红的类别数量: {len(red_categories)}")

        # 第二遍：写入数据并应用格式
        for file_name, result_info in results.items():
            try:
                # 获取文件索引（重复上面的逻辑）
                if is_combined:
                    parts = file_name.split('_')
                    if len(parts) >= 3 and parts[-1].endswith('.png'):
                        index_str = parts[-1].split('.')[0]
                        if index_str.isdigit():
                            index = int(index_str)
                        else:
                            continue
                    else:
                        continue
                else:
                    if not file_name.startswith("cell_"):
                        continue
                    index = int(file_name.split("_")[1].split(".")[0])

                # 计算行列位置
                row = ((index - 1) // 16) + 1
                col = ((index - 1) % 16) + 1

                # 提取预测标签和置信度
                if isinstance(result_info, dict):
                    digit = result_info.get('label', 'Unknown')
                    confidence = result_info.get('confidence', 1.0)
                    display_text = f"{digit}({confidence:.2f})"
                else:
                    digit = result_info
                    display_text = str(digit)
                    confidence = 1.0

                # 写入 Excel 单元格
                cell = ws.cell(row=row, column=col, value=display_text)
                
                # 确定当前单元格所属的类别
                current_category = None
                if col == 1:
                    current_category = None  # 第一列不属于任何类别
                elif 2 <= col <= 6:
                    current_category = 1
                elif 7 <= col <= 11:
                    current_category = 2
                elif 12 <= col <= 16:
                    current_category = 3
                
                # 设置字体颜色和背景色
                if current_category and (row, current_category) in red_categories:
                    # 该类别需要标红：红色背景，白色字体
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    # print(f"应用红色背景到单元格 ({row}, {col}) - 类别{current_category}")  # 移除这行
                else:
                    # 正常格式：根据置信度设置字体颜色
                    font_color = "FF0000" if confidence < 0.65 else "000000"
                    cell.font = Font(color=font_color)

            except (IndexError, ValueError) as e:
                print(f"写入文件 {file_name} 时出错: {e}")
                continue

        # 设置列宽和居中对齐
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if cell.value:
                    try:
                        cell_length = 0
                        for char in str(cell.value):
                            if ord(char) > 127:
                                cell_length += 2
                            else:
                                cell_length += 1
                        if cell_length > max_length:
                            max_length = cell_length
                    except:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))

            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = adjusted_width

        # 设置单元格居中
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 保存 Excel 文件
        wb.save(excel_path)
        print(f"预测结果已保存到 {excel_path}")
        
    except Exception as e:
        print(f"创建Excel文件时出错: {e}")
        import traceback
        traceback.print_exc()


def visualize_results(results, skipped_files, input_folder, output_path):
    """
    可视化预测结果，包括成功预测和被跳过的图片，使用固定的16列布局
    :param results: 预测结果字典 {文件名: 预测值}
    :param skipped_files: 被跳过的文件字典 {文件名: 跳过原因}
    :param input_folder: 输入图片文件夹路径
    :param output_path: 输出可视化图片路径
    """
    # 合并所有文件，包括成功预测和跳过的
    all_files = list(results.keys()) + list(skipped_files.keys())

    # 定义文件排序函数
    def get_file_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')

    # 按数字顺序排序所有文件
    all_files.sort(key=get_file_index)

    n_samples = len(all_files)
    if n_samples == 0:
        print("没有图片可视化")
        return

    # 固定为16列布局
    fixed_cols = 16
    rows = (n_samples + fixed_cols - 1) // fixed_cols

    # 创建画布，设置更大的尺寸以适应16列
    fig_width = fixed_cols * 1.2  # 每列1.2英寸宽
    fig_height = rows * 1.2  # 每行1.2英寸高
    fig, axes = plt.subplots(rows, fixed_cols, figsize=(fig_width, fig_height))

    # 确保axes始终是二维数组
    if rows == 1:
        axes = axes.reshape(1, -1)

    # 绘制每个图片，按行列位置排列
    for i, file_name in enumerate(all_files):
        # 将单元格索引转换为行列位置
        file_index = get_file_index(file_name)
        if file_index == float('inf'):  # 对于无法解析索引的文件
            row_idx = i // fixed_cols
            col_idx = i % fixed_cols
        else:
            # 修正：索引从1开始，需要先减1
            row_idx = (file_index - 1) // fixed_cols
            col_idx = (file_index - 1) % fixed_cols

        # 确保不超出画布范围
        if row_idx >= rows or col_idx >= fixed_cols:
            print(f"警告: 文件 {file_name} (索引: {file_index}) 超出画布范围")
            continue

        # 读取图像
        image_path = os.path.join(input_folder, file_name)
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)

        if image is None:
            # 处理无法读取的图片
            axes[row_idx, col_idx].text(0.5, 0.5, "图片加载失败",
                                        ha='center', va='center', fontsize=7)
            axes[row_idx, col_idx].set_title(f"#{file_index}", fontsize=7)
        else:
            # 显示图像
            axes[row_idx, col_idx].imshow(image, cmap='gray')

                    # 设置标题
        if file_name in results:
            result_info = results[file_name]
            if isinstance(result_info, dict):
                label = result_info['label']
                confidence = result_info['confidence']
                title_text = f"#{file_index}: {label}({confidence:.2f})"
                # 根据置信度设置颜色：低于0.65为红色，否则为蓝色
                title_color = 'red' if confidence < 0.65 else 'blue'
            else:
                # 向后兼容旧格式
                title_text = f"#{file_index}: {result_info}"
                title_color = 'blue'
            axes[row_idx, col_idx].set_title(title_text, fontsize=7, color=title_color)
        else:
            reason = skipped_files.get(file_name, "未知原因")
            # 使用简短的原因说明，以适应小尺寸标题
            short_reason = "黑" if reason == "接近纯黑" else "错误"
            axes[row_idx, col_idx].set_title(f"#{file_index}: {short_reason}",
                                             fontsize=7, color='red')

        axes[row_idx, col_idx].axis('off')

    # 隐藏未使用的子图
    for r in range(rows):
        for c in range(fixed_cols):
            # 检查此位置是否有对应文件显示
            has_image = False
            for file_name in all_files:
                file_index = get_file_index(file_name)
                if file_index != float('inf'):
                    if file_index // fixed_cols == r and file_index % fixed_cols == c:
                        has_image = True
                        break

            # 若无图像显示，则隐藏坐标轴
            if not has_image:
                axes[r, c].axis('off')

    plt.tight_layout()
    plt.subplots_adjust(hspace=0.4)  # 增加行间距，避免标题重叠
    plt.savefig(output_path, dpi=150)  # 使用更高的DPI获得更清晰的图像
    plt.close()  # 关闭图形以释放内存
    print(f"可视化结果(16列布局)已保存到 {output_path}")



def print_detailed_predictions(results, skipped_files):
    """
    显示简要的预测结果统计
    """
    if not results and not skipped_files:
        return
    
    # 统计置信度分布
    if results:
        high_count = medium_count = low_count = 0
        for file_name, result_info in results.items():
            if isinstance(result_info, dict):
                confidence = result_info['confidence']
                if confidence > 0.9:
                    high_count += 1
                elif confidence > 0.7:
                    medium_count += 1
                else:
                    low_count += 1
        
        print(f"预测完成: 高置信度({high_count}) 中等置信度({medium_count}) 低置信度({low_count})")
    
    if skipped_files:
        print(f"跳过文件: {len(skipped_files)}个")


def detect_chinese_in_cropped_table(cropped_cell, img_cell_dir):
    """
    检测YOLO裁剪后的完整表格区域中是否包含手写汉字内容
    专门针对手写汉字优化的检测算法
    :param cropped_cell: YOLO检测后的表格区域图像(numpy数组)
    :param img_cell_dir: 输出目录
    返回: (是否包含汉字, 检测结果字典)
    """
    # 检查是否启用汉字检测
    if not CHINESE_DETECTION_CONFIG.get('enable_chinese_detection', True):
        return False, {'message': '汉字检测已禁用'}
    
    chinese_detected = False
    detection_results = {}
    
    try:
        # cropped_cell 是numpy数组，转换为灰度图
        if len(cropped_cell.shape) == 3:
            gray = cv2.cvtColor(cropped_cell, cv2.COLOR_BGR2GRAY)
        else:
            gray = cropped_cell.copy()
        
        # 检查图像尺寸
        h, w = gray.shape
        if h < 50 or w < 50:
            return False, {'message': '表格区域太小，跳过汉字检测'}
        
        config = CHINESE_DETECTION_CONFIG
        chinese_indicators = 0
        
        # === 手写汉字检测算法 ===
        
        # 1. 自适应二值化（对手写字体更敏感）
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
        
        # 2. 检测特征1：连通域分析（手写汉字通常有多个分离的笔画）
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        
        # 分析有效的字符级连通域（加强过滤条件）
        valid_components = []
        for contour in contours:
            area = cv2.contourArea(contour)
            x, y, w, h = cv2.boundingRect(contour)
            
            # 更严格的手写汉字判断条件
            if (area > 200 and  # 提高最小面积要求，汉字通常较大
                min(w, h) > 15 and  # 提高最小尺寸，避免小噪点
                max(w, h) < min(gray.shape) * 0.6 and  # 不能太大
                0.4 <= w/h <= 2.5):  # 汉字通常接近方形
                
                # 计算更多特征来区分汉字和数字
                hull = cv2.convexHull(contour)
                hull_area = cv2.contourArea(hull)
                solidity = area / hull_area if hull_area > 0 else 0
                
                # 计算轮廓复杂度（汉字笔画更复杂）
                perimeter = cv2.arcLength(contour, True)
                compactness = perimeter * perimeter / (4 * np.pi * area) if area > 0 else 0
                
                # 计算凹陷数量（汉字通常有更多凹陷）
                hull_defects = len(cv2.convexityDefects(contour, cv2.convexHull(contour, returnPoints=False))) if len(contour) > 3 else 0
                
                # 手写汉字特征过滤
                is_likely_hanzi = (
                    solidity < 0.9 and  # 汉字通常不是完全实心的
                    compactness > 2.0 and  # 汉字边界相对复杂
                    area > 300 and  # 汉字面积通常较大
                    min(w, h) > 20  # 汉字尺寸通常较大
                )
                
                if is_likely_hanzi:
                    valid_components.append({
                        'contour': contour,
                        'area': area,
                        'bbox': (x, y, w, h),
                        'solidity': solidity,
                        'aspect_ratio': w/h,
                        'compactness': compactness,
                        'hull_defects': hull_defects
                    })
        
        # 指标1：有效字符数量（提高要求，减少误检）
        min_components = max(1, config.get('min_char_components', 2))  # 至少需要1个高质量汉字
        if len(valid_components) >= min_components:
            chinese_indicators += 1
            print(f"检测到 {len(valid_components)} 个有效汉字区域")
        else:
            print(f"有效汉字区域不足: {len(valid_components)}/{min_components}")
        
        # 3. 检测特征2：笔画密度分析（调整范围，减少误检）
        # 计算前景像素密度
        total_foreground = np.sum(binary == 255)
        total_pixels = binary.size
        foreground_ratio = total_foreground / total_pixels
        
        # 汉字的前景密度通常在特定范围内
        min_fg = config.get('min_foreground_ratio', 0.08)  # 提高下限
        max_fg = config.get('max_foreground_ratio', 0.25)  # 降低上限，避免数字表格
        if min_fg <= foreground_ratio <= max_fg:
            chinese_indicators += 1
            print(f"前景密度: {foreground_ratio:.3f} (符合汉字特征)")
        else:
            print(f"前景密度不符合: {foreground_ratio:.3f} (期望: {min_fg}-{max_fg})")
        
        # 4. 检测特征3：边缘复杂度（提高阈值）
        edges = cv2.Canny(gray, 50, 150)
        edge_density = np.sum(edges == 255) / edges.size
        
        # 汉字边缘应该相当复杂
        min_edge = config.get('min_edge_density', 0.05)  # 提高阈值
        if edge_density > min_edge:
            chinese_indicators += 1
            print(f"边缘密度: {edge_density:.3f} (符合汉字特征)")
        else:
            print(f"边缘密度不足: {edge_density:.3f} (期望: >{min_edge})")
        
        # 5. 检测特征4：局部复杂度分析（手写汉字局部变化丰富）
        # 使用局部二值模式检测纹理复杂度
        def calculate_local_complexity(img, block_size=30):
            """计算图像的局部复杂度"""
            complexity_scores = []
            for y in range(0, img.shape[0] - block_size, block_size // 2):
                for x in range(0, img.shape[1] - block_size, block_size // 2):
                    block = img[y:y+block_size, x:x+block_size]
                    
                    # 计算块内的梯度变化
                    grad_x = cv2.Sobel(block, cv2.CV_64F, 1, 0, ksize=3)
                    grad_y = cv2.Sobel(block, cv2.CV_64F, 0, 1, ksize=3)
                    magnitude = np.sqrt(grad_x**2 + grad_y**2)
                    
                    # 计算梯度的标准差（复杂度指标）
                    if magnitude.size > 0:
                        complexity = np.std(magnitude)
                        complexity_scores.append(complexity)
            
            return np.mean(complexity_scores) if complexity_scores else 0
        
        local_complexity = calculate_local_complexity(gray)
        min_complexity = config.get('min_local_complexity', 15)
        if local_complexity > min_complexity:
            chinese_indicators += 1
            print(f"局部复杂度: {local_complexity:.2f} (符合汉字特征)")
        else:
            print(f"局部复杂度不足: {local_complexity:.2f} (期望: >{min_complexity})")
        
        # 6. 检测特征5：字符分布分析（手写汉字通常分布在多个区域）
        dist_std = 0
        if len(valid_components) > 0:
            # 计算字符间的空间分布
            centers = []
            for comp in valid_components:
                x, y, w, h = comp['bbox']
                centers.append((x + w//2, y + h//2))
            
            if len(centers) >= 2:
                # 计算字符间距离的变异性
                distances = []
                for i in range(len(centers)):
                    for j in range(i+1, len(centers)):
                        dist = np.sqrt((centers[i][0] - centers[j][0])**2 + (centers[i][1] - centers[j][1])**2)
                        distances.append(dist)
                
                if distances:
                    dist_std = np.std(distances)
                    min_dist_std = config.get('min_distance_std', 30)
                    # 手写汉字字符间距离有一定变化
                    if dist_std > min_dist_std:
                        chinese_indicators += 1
                        print(f"字符分布变异性: {dist_std:.2f} (符合汉字特征)")
                    else:
                        print(f"字符分布变异性不足: {dist_std:.2f} (期望: >{min_dist_std})")
        
        # 综合判断
        min_indicators_required = config.get('min_indicators', 1)
        chinese_detected = chinese_indicators >= min_indicators_required
        
        detection_results = {
            'chinese_indicators': chinese_indicators,
            'valid_components_count': len(valid_components),
            'foreground_ratio': total_foreground / total_pixels,
            'edge_density': edge_density,
            'local_complexity': local_complexity,
            'distance_std': dist_std,
            'detection_confidence': chinese_indicators / 5.0,  # 5个指标的置信度
            'min_indicators_required': min_indicators_required
        }
        
        # 创建检测结果可视化图像（显示所有连通域，用于调试）
        debug_img = cropped_cell.copy()
        if len(debug_img.shape) == 2:
            debug_img = cv2.cvtColor(debug_img, cv2.COLOR_GRAY2BGR)
        
        # 首先显示所有原始连通域（蓝色）
        all_components = []
        for i, contour in enumerate(contours):
            area = cv2.contourArea(contour)
            if area > 50:  # 只显示面积大于50的区域
                x, y, w, h = cv2.boundingRect(contour)
                all_components.append({
                    'contour': contour,
                    'bbox': (x, y, w, h),
                    'area': area,
                    'aspect_ratio': w/h if h > 0 else 0,
                    'type': 'all'
                })
        
        # 绘制所有连通域（蓝色）
        for i, comp in enumerate(all_components):
            x, y, w, h = comp['bbox']
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (255, 0, 0), 1)  # 蓝色细线
            cv2.putText(debug_img, f"A{i+1}", (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (255, 0, 0), 1)
        
        # 绘制通过基本筛选的字符区域（橙色）
        for i, comp in enumerate(valid_components):
            x, y, w, h = comp['bbox']
            cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 165, 255), 2)  # 橙色粗线
            
            # 添加详细信息
            area = comp['area']
            aspect_ratio = comp['aspect_ratio']
            solidity = comp.get('solidity', 0)
            cv2.putText(debug_img, f"V{i+1}", (x, y-5), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 165, 255), 1)
            cv2.putText(debug_img, f"A:{int(area)}", (x, y+h+15), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 165, 255), 1)
            cv2.putText(debug_img, f"AR:{aspect_ratio:.1f}", (x, y+h+30), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 165, 255), 1)
            cv2.putText(debug_img, f"S:{solidity:.2f}", (x, y+h+45), cv2.FONT_HERSHEY_SIMPLEX, 0.3, (0, 165, 255), 1)
        
        # 如果检测到汉字，用绿色突出显示
        if chinese_detected and len(valid_components) > 0:
            for i, comp in enumerate(valid_components):
                x, y, w, h = comp['bbox']
                cv2.rectangle(debug_img, (x, y), (x + w, y + h), (0, 255, 0), 3)  # 绿色超粗线
        
        # 在图像上添加检测统计信息
        info_text = [
            f"Total Contours: {len(all_components)}",
            f"Valid Components: {len(valid_components)}",
            f"FG Density: {total_foreground/total_pixels:.3f}",
            f"Edge Density: {edge_density:.3f}",
            f"Local Complex: {local_complexity:.1f}",
            f"Distance Std: {dist_std:.1f}",
            f"Indicators: {chinese_indicators}/5",
            f"Chinese: {'YES' if chinese_detected else 'NO'}"
        ]
        
        # 添加颜色说明
        legend_text = [
            "Legend:",
            "Blue (A#): All contours",
            "Orange (V#): Valid components", 
            "Green: Chinese detected"
        ]
        
        # 显示统计信息（左上角）
        for i, text in enumerate(info_text):
            y_pos = 20 + i * 18
            # 白色底，黑色字
            cv2.putText(debug_img, text, (11, y_pos+1), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 2)
            cv2.putText(debug_img, text, (10, y_pos), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        
        # 显示图例（右上角）
        img_width = debug_img.shape[1]
        for i, text in enumerate(legend_text):
            y_pos = 20 + i * 18
            x_pos = img_width - 200
            cv2.putText(debug_img, text, (x_pos+1, y_pos+1), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (255, 255, 255), 2)
            cv2.putText(debug_img, text, (x_pos, y_pos), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 0, 0), 1)
        
        # 保存检测结果图像
        detection_image_path = os.path.join(img_cell_dir, "chinese_character_detection.png")
        cv2.imwrite(detection_image_path, debug_img)
        print(f"汉字检测调试图已保存: {detection_image_path}")
        
        if chinese_detected:
            confidence = detection_results['detection_confidence']
            char_count = detection_results['valid_components_count']
            print(f"⚠️ 检测到手写汉字表格 (置信度: {confidence:.2f}, 满足指标: {chinese_indicators}/{min_indicators_required})")
            print(f"   有效字符区域: {char_count}个, 前景密度: {detection_results['foreground_ratio']:.3f}")
        else:
            print(f"未检测到汉字 (满足指标: {chinese_indicators}/{min_indicators_required})")
        
        return chinese_detected, detection_results
        
    except Exception as e:
        print(f"表格区域汉字检测失败: {e}")
        return False, {}


# 旧的复杂汉字检测算法已删除，将用新的简单算法替代


def create_chinese_cells_collage(chinese_cells_info, cell_folder, output_path, confidence_threshold=0.6):
    """
    将检测到汉字的单元格拼接成一张图片
    :param chinese_cells_info: 汉字单元格信息列表
    :param cell_folder: 单元格文件夹路径
    :param output_path: 输出图片路径
    :param confidence_threshold: 置信度阈值，只显示高于此值的单元格
    """
    # 过滤高置信度的汉字单元格
    high_confidence_cells = [cell for cell in chinese_cells_info if cell['confidence'] >= confidence_threshold]
    
    if not high_confidence_cells:
        print(f"没有置信度大于{confidence_threshold}的汉字单元格")
        return 0
    
    print(f"拼接 {len(high_confidence_cells)} 个高置信度汉字单元格（阈值>{confidence_threshold}）")
        
    # 按置信度降序排列
    high_confidence_cells.sort(key=lambda x: x['confidence'], reverse=True)
    
    # 设置拼接参数
    cell_size = 120  # 每个单元格显示大小
    cols_per_row = 6  # 每行显示6个单元格
    padding = 10  # 单元格间距
    
    # 计算网格尺寸
    import math
    rows = math.ceil(len(high_confidence_cells) / cols_per_row)
    
    # 计算画布尺寸
    canvas_width = cols_per_row * cell_size + (cols_per_row + 1) * padding
    title_height = 60  # 标题区域高度
    canvas_height = rows * (cell_size + 40) + (rows + 1) * padding + title_height  # 40是文字区域高度
    
    # 创建白色画布
    canvas = np.ones((canvas_height, canvas_width, 3), dtype=np.uint8) * 255
    
    # 添加标题
    title_text = f"检测到的汉字单元格 (置信度 > {confidence_threshold})"
    font = cv2.FONT_HERSHEY_SIMPLEX
    title_font_scale = 1.0
    title_thickness = 2
    title_size = cv2.getTextSize(title_text, font, title_font_scale, title_thickness)[0]
    title_x = (canvas_width - title_size[0]) // 2
    title_y = 35
    
    cv2.putText(canvas, title_text, (title_x, title_y), font, title_font_scale, (0, 0, 0), title_thickness)
    
    # 拼接每个单元格
    for i, cell_info in enumerate(high_confidence_cells):
        row = i // cols_per_row
        col = i % cols_per_row
        
        # 计算位置
        x = col * (cell_size + padding) + padding
        y = row * (cell_size + 40 + padding) + padding + title_height
        
        # 读取单元格图片
        cell_path = os.path.join(cell_folder, cell_info['file_name'])
        cell_image = cv2.imread(cell_path)
        
        if cell_image is not None:
            # 调整图片大小
            cell_resized = cv2.resize(cell_image, (cell_size, cell_size))
            
            # 放置到画布上
            canvas[y:y+cell_size, x:x+cell_size] = cell_resized
            
            # 添加边框
            cv2.rectangle(canvas, (x-1, y-1), (x+cell_size+1, y+cell_size+1), (0, 0, 0), 2)
            
            # 添加单元格信息
            info_text = f"#{cell_info['cell_index']} ({cell_info['confidence']:.2f})"
            info_y = y + cell_size + 20
            info_size = cv2.getTextSize(info_text, font, 0.5, 1)[0]
            info_x = x + (cell_size - info_size[0]) // 2
            
            # 白色背景
            cv2.rectangle(canvas, (info_x-2, info_y-15), (info_x+info_size[0]+2, info_y+5), (255, 255, 255), -1)
            # 黑色文字
            cv2.putText(canvas, info_text, (info_x, info_y), font, 0.5, (0, 0, 0), 1)
            
            # 根据置信度设置颜色提示
            if cell_info['confidence'] >= 0.8:
                color = (0, 255, 0)  # 绿色 - 高置信度
            elif cell_info['confidence'] >= 0.7:
                color = (0, 165, 255)  # 橙色 - 中等置信度
            else:
                color = (0, 0, 255)  # 红色 - 低置信度
            
            # 在右上角添加彩色置信度指示
            cv2.circle(canvas, (x + cell_size - 10, y + 10), 5, color, -1)
    
    # 添加图例
    legend_y = canvas_height - 30
    legend_items = [
        ("● 高置信度 (≥0.8)", (0, 255, 0)),
        ("● 中等置信度 (≥0.7)", (0, 165, 255)),
        ("● 低置信度 (≥0.6)", (0, 0, 255))
    ]
    
    legend_x = 20
    for text, color in legend_items:
        cv2.putText(canvas, text, (legend_x, legend_y), font, 0.4, color, 1)
        legend_x += 160
    
    # 保存拼接图片
    cv2.imwrite(output_path, canvas)
    print(f"汉字单元格拼接图已保存: {output_path}")
    
    return len(high_confidence_cells)


def detect_chinese_by_mnist_reverse(processed_cell_folder, model, transform, exclude_rows=None):
    """
    基于MNIST模型的反向汉字检测 - 更简单、更可靠的方法
    
    核心思路：
    1. 用训练好的MNIST模型预测每个单元格
    2. 如果模型置信度很低，说明不是数字/字母
    3. 再检查是否有足够的内容，有内容就认为是汉字
    
    :param processed_cell_folder: 预处理后的单元格文件夹路径
    :param model: 训练好的MNIST模型
    :param transform: 图像预处理变换
    :param exclude_rows: 要排除的行号列表
    :return: (是否检测到汉字, 汉字单元格列表)
    """
    if exclude_rows is None:
        exclude_rows = [2]  # 默认排除第2行
    
    chinese_cells = []
    
    # 获取所有单元格文件
    cell_files = []
    for file_name in os.listdir(processed_cell_folder):
        if file_name.startswith('cell_') and file_name.endswith('.png') and file_name != 'cell_0.png':
            cell_files.append(file_name)
    
    # 按数字排序
    def get_cell_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')
    
    cell_files.sort(key=get_cell_index)
    
    print(f"开始MNIST反向汉字检测，共{len(cell_files)}个单元格")
    
    # 检测每个单元格
    for file_name in cell_files:
        file_path = os.path.join(processed_cell_folder, file_name)
        cell_index = get_cell_index(file_name)
        
        # 计算单元格所在的行（假设16列布局）
        cols_per_row = 16
        row_number = ((cell_index - 1) // cols_per_row) + 1
        
        # 检查是否在排除行中
        if row_number in exclude_rows:
            continue
            
        # 读取并检查图像
        image = cv2.imread(file_path)
        if image is None:
            continue
        
        # 转换为灰度图
        if len(image.shape) == 3:
            image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # 确保数据类型正确
        image = image.astype(np.uint8)
        
        # 1. 基础内容检查：过滤明显的空白格
        total_pixels = image.size
        # 对于预处理后的图像，白色是前景
        white_pixels = np.sum(image > 200)  # 接近白色的像素
        content_ratio = white_pixels / total_pixels
        
        if content_ratio < 0.02:  # 内容太少，明显是空白格
            continue
        
        if content_ratio > 0.8:  # 内容太多，可能是噪声或错误
            continue
        
        # 2. 使用MNIST模型预测
        try:
            
            # 转换为PIL图像并预处理
            pil_image = Image.fromarray(image, mode='L')  # 明确指定为灰度图
            input_tensor = transform(pil_image).unsqueeze(0)
            
            # 预测
            with torch.no_grad():
                output = model(input_tensor)
                probabilities = torch.softmax(output, dim=1)
                max_confidence = torch.max(probabilities).item()
                predicted_idx = torch.argmax(probabilities).item()
            
            # 3. 反向逻辑判断
            # 如果MNIST模型的最高置信度很低，说明可能不是数字/字母
            mnist_threshold = 0.4  # MNIST置信度阈值，可调
            
            if max_confidence < mnist_threshold:
                # 再次检查是否有足够内容（避免噪点）
                if content_ratio > 0.05:  # 有一定内容
                    # 额外的汉字特征检查（简单版）
                    is_likely_chinese = simple_chinese_check(image)
                    if is_likely_chinese:
                        # 反向置信度：MNIST越不确定，我们越确定是汉字
                        confidence = 0.9 - max_confidence  # 反向置信度
                        chinese_cells.append({
                            'file_name': file_name,
                            'cell_index': cell_index,
                            'row': row_number,
                            'confidence': min(confidence, 0.95),  # 最高95%
                            'mnist_confidence': max_confidence,
                            'content_ratio': content_ratio
                        })
                        print(f"MNIST反向检测到汉字: {file_name} (第{row_number}行, MNIST置信度:{max_confidence:.2f}, 汉字置信度:{confidence:.2f})")
            
        except Exception as e:
            print(f"MNIST检测失败 {file_name}: {e}")
            continue
    
    has_chinese = len(chinese_cells) > 0
    if has_chinese:
        print(f"MNIST反向检测到 {len(chinese_cells)} 个可能的汉字单元格")
    
    return has_chinese, chinese_cells


def simple_chinese_check(image):
    """
    简单的汉字特征检查，避免复杂的特征工程
    
    :param image: 灰度图像（预处理后的28x28图像）
    :return: 是否可能是汉字
    """
    try:
        # 确保图像是正确的格式
        if len(image.shape) == 3:
            image = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        
        # 确保数据类型正确
        image = image.astype(np.uint8)
        
        # 对于预处理后的图像，需要反转：白色是前景，黑色是背景
        # 转换为前景为255，背景为0的二值图像
        _, binary = cv2.threshold(image, 127, 255, cv2.THRESH_BINARY)
        
        # 1. 检查连通域数量和复杂度
        contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        valid_contours = [cnt for cnt in contours if cv2.contourArea(cnt) > 5]  # 降低面积要求
        
        if len(valid_contours) == 0:
            return False
        
        # 2. 计算总的前景像素比例
        foreground_pixels = np.sum(binary == 255)
        total_pixels = binary.size
        fg_ratio = foreground_pixels / total_pixels
        
        # 前景密度合理检查
        if fg_ratio < 0.05 or fg_ratio > 0.7:
            return False
        
        # 3. 检查最大连通域的特征
        largest_contour = max(valid_contours, key=cv2.contourArea)
        area = cv2.contourArea(largest_contour)
        
        if area < 10:  # 太小
            return False
        
        # 4. 计算宽高比
        x, y, w, h = cv2.boundingRect(largest_contour)
        aspect_ratio = w / h if h > 0 else 0
        
        # 5. 计算复杂度
        perimeter = cv2.arcLength(largest_contour, True)
        compactness = perimeter / np.sqrt(area) if area > 0 else 0
        
        # 6. 简单判断规则：
        # 情况1：有多个连通域（可能是汉字的多个笔画）
        if len(valid_contours) >= 2:
            return True
        
        # 情况2：单个连通域但比较复杂且宽高比合理
        elif len(valid_contours) == 1:
            # 复杂度合理 && 宽高比不是极端值 && 有足够的前景
            if (compactness > 3.5 and 
                0.3 <= aspect_ratio <= 3.0 and 
                fg_ratio > 0.1):
                return True
        
        return False
        
    except Exception:
        return False


def detect_chinese_in_cells(cell_folder, exclude_rows=None):
    """
    兼容性包装函数：使用MNIST反向检测
    """
    # 为了兼容现有代码，我们需要先进行预处理，然后使用MNIST模型
    # 但这需要模型和transform，暂时返回空结果
    print("警告：旧的汉字检测函数已被禁用，请使用MNIST反向检测")
    return False, []





def check_severe_distortion(img_cell_dir):
    """
    检查是否存在严重扭曲的垂直线、异常的水平线间距
    注意: 汉字检测已在主函数的 cropped_cell 阶段完成
    返回: (是否有严重扭曲, 分析结果字典)
    """
    distortion_issues = []
    analysis_results = {}
    filter_reasons = []  # 新增：记录具体过滤原因
    
    # 1. 检查垂直线弯曲度
    vertical_img_path = os.path.join(img_cell_dir, "step2_vertical_lines_original.png")
    if os.path.exists(vertical_img_path):
        try:
            lines_analysis = analyze_vertical_lines_curvature(
                vertical_img_path, 
                img_cell_dir, 
                min_length_threshold=150,
                verbose=False
            )
            
            # 检查是否有状态码为"3"的严重弯曲线条
            severe_curved_lines = [line for line in lines_analysis if line.get('curvature_code') == "3"]
            if severe_curved_lines:
                issue_msg = f"检测到{len(severe_curved_lines)}条严重弯曲的垂直线"
                distortion_issues.append(issue_msg)
                filter_reasons.append("垂直线严重弯曲")  # 新增
                analysis_results['vertical_distortion'] = {
                    'severe_count': len(severe_curved_lines),
                    'total_lines': len(lines_analysis)
                }
            
        except Exception as e:
            print(f"垂直线分析失败: {e}")
    
    # 2. 检查水平线间距异常
    horizontal_img_path = os.path.join(img_cell_dir, "step2_horizontal_lines.png")
    if os.path.exists(horizontal_img_path):
        try:
            horizontal_img = cv2.imread(horizontal_img_path, cv2.IMREAD_GRAYSCALE)
            spacing_analysis = analyze_horizontal_lines_spacing(horizontal_img, img_cell_dir, verbose=False)
            
            if spacing_analysis and spacing_analysis.get('has_small_spacing', False):
                small_distances = [d for d in spacing_analysis['distances'] if d < 35]
                # 只有确实检测到小间距才添加问题
                if len(small_distances) > 0:
                    issue_msg = f"检测到{len(small_distances)}个小于35像素的水平线间距"
                distortion_issues.append(issue_msg)
                filter_reasons.append("水平线间距过小")  # 新增
                analysis_results['horizontal_spacing'] = {
                    'small_distances': small_distances,
                    'min_distance': spacing_analysis['min_distance']
                }
                
        except Exception as e:
            print(f"水平线间距分析失败: {e}")
    
    # 3. 检查表格区域中是否包含汉字内容（新增）
    # 注意：这个检测需要在主函数中的 cropped_cell 阶段进行
    # 这里暂时跳过，实际检测在主函数中完成
    pass
    
    # 判断是否有严重扭曲
    has_severe_distortion = len(distortion_issues) > 0
    
    if has_severe_distortion:
        print(f"扭曲问题: {'; '.join(distortion_issues)}")
        print(f"过滤原因: {' + '.join(filter_reasons)}")  # 新增：明确显示过滤原因
    
    analysis_results['issues'] = distortion_issues
    analysis_results['filter_reasons'] = filter_reasons  # 新增
    
    return has_severe_distortion, analysis_results


def save_distorted_files(img_name, image_path, img_excel_path, img_vis_path, 
                        img_cell_dir, bad_timestamp_dir, filter_reasons=None):
    """
    保存严重扭曲的图片相关文件到待人工处理文件夹
    """
    try:
        # 直接使用时间戳目录，不创建图片子文件夹
        os.makedirs(bad_timestamp_dir, exist_ok=True)
        
        # 保存过滤原因到全局记录中
        if not hasattr(save_distorted_files, 'filtered_records'):
            save_distorted_files.filtered_records = []
        
        save_distorted_files.filtered_records.append({
            'img_name': img_name,
            'filter_reasons': filter_reasons or [],
            'filter_time': time.strftime('%Y-%m-%d %H:%M:%S')
        })
        
        # 移除单独的过滤原因文件创建
        # if filter_reasons:
        #     reason_file = os.path.join(bad_timestamp_dir, f"{img_name}_filter_reason.txt")
        #     with open(reason_file, 'w', encoding='utf-8') as f:
        #         f.write(f"图片: {img_name}\n")
        #         f.write(f"过滤原因: {' + '.join(filter_reasons)}\n")
        #         f.write(f"过滤时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
        
        # 1. 复制原始图片
        original_dest = os.path.join(bad_timestamp_dir, f"{img_name}_original.png")
        original_img = cv2.imread(image_path)
        if original_img is not None:
            cv2.imwrite(original_dest, original_img)
            print(f"已保存原始图片: {original_dest}")
        
        # 2. 复制Excel文件
        if os.path.exists(img_excel_path):
            excel_dest = os.path.join(bad_timestamp_dir, f"{img_name}_predictions.xlsx")
            import shutil
            shutil.copy2(img_excel_path, excel_dest)
            print(f"已保存Excel文件: {excel_dest}")
        
        # 3. 复制垂直线分析图
        analysis_file = os.path.join(img_cell_dir, "long_vertical_lines_curvature_analysis.png")
        if os.path.exists(analysis_file):
            analysis_dest = os.path.join(bad_timestamp_dir, f"{img_name}_vertical_lines_analysis.png")
            shutil.copy2(analysis_file, analysis_dest)
            print(f"已保存分析图: {analysis_dest}")
        
        # 4. 复制检测结果可视化图
        if os.path.exists(img_vis_path):
            vis_dest = os.path.join(bad_timestamp_dir, f"{img_name}_detection_results.png")
            shutil.copy2(img_vis_path, vis_dest)
            print(f"已保存可视化图: {vis_dest}")
        
        return True
        
    except Exception as e:
        print(f"保存扭曲文件失败 {img_name}: {e}")
        return False


def main(input_path, yolo_model_path):
    """
    主函数，整合各个处理步骤
    当检测到严重扭曲时，保存到待人工处理文件夹
    """
    print("开始处理OCR任务")
    
    start_time = time.time()

    # 1. 创建带时间戳的文件夹
    folders = create_timestamped_folders()
    
    # 2. 创建待人工处理文件夹的时间戳目录
    bad_base_dir = r"D:\OCR-project\weld_report\待人工处理"
    timestamp = folders["timestamp"]  # 使用相同的时间戳
    bad_timestamp_dir = os.path.join(bad_base_dir, timestamp)
    os.makedirs(bad_timestamp_dir, exist_ok=True)

    # 3. 创建待纠偏处理文件夹的时间戳目录
    correction_base_dir = r"D:\OCR-project\weld_report\待纠偏处理"
    correction_timestamp_dir = os.path.join(correction_base_dir, timestamp)
    os.makedirs(correction_timestamp_dir, exist_ok=True)

    # 处理图片路径
    image_paths = []
    if os.path.isdir(input_path):
        for file_name in os.listdir(input_path):
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                image_paths.append(os.path.join(input_path, file_name))
        print(f"找到 {len(image_paths)} 张图片")
    else:
        image_paths = [input_path]

    if not image_paths:
        print("未找到任何图片")
        return

    # 统计变量
    total_processed = 0
    vertical_distorted_count = 0      # 垂直线扭曲计数
    horizontal_distorted_count = 0    # 水平线间距问题计数
    both_distorted_count = 0          # 两种问题都有的计数
    chinese_detected_count = 0        # 新增：包含汉字的图片计数
    
    # 初始化过滤记录
    if hasattr(save_distorted_files, 'filtered_records'):
        delattr(save_distorted_files, 'filtered_records')

    # 3. 处理每张图片
    for idx, image_path in enumerate(image_paths):
        print(f"\n处理图片 {idx + 1}/{len(image_paths)}: {os.path.basename(image_path)}")

        # 为每张图片创建子文件夹
        img_name = os.path.splitext(os.path.basename(image_path))[0]
        img_cell_dir = os.path.join(folders['cell_dir'], img_name)
        img_processed_dir = os.path.join(folders['processed_cells_dir'], img_name)
        os.makedirs(img_cell_dir, exist_ok=True)
        os.makedirs(img_processed_dir, exist_ok=True)

        # 步骤1: YOLO检测并分割单元格
        cropped_cell = detect_and_crop_largest_table_cell(
            image_path,
            output_folder=img_cell_dir,
            model_path=yolo_model_path,
            min_cell_size=20,
            conf_threshold=0.2
        )

        if cropped_cell is None:
            print(f"未检测到表格: {img_name}")
            continue

        # 保存裁剪后的表格
        cv2.imwrite(os.path.join(img_cell_dir, "cropped_cell.png"), cropped_cell)

        # 检查汉字检测配置
        config = CHINESE_DETECTION_CONFIG
        detection_level = config.get('detection_level', 'table')
        
        # 如果配置为表格级检测，执行原有逻辑
        if detection_level == 'table' and config.get('enable_chinese_detection', False):
            has_chinese_in_table, chinese_detection_results = detect_chinese_in_cropped_table(cropped_cell, img_cell_dir)
            
            if has_chinese_in_table:
                print(f"⚠️ 检测到表格包含汉字，进行单元格分割并生成拼接图")
                
                # 先分割单元格，然后进行单元格级汉字检测
                small_cells = split_into_small_cells(cropped_cell, output_folder=img_cell_dir)
                
                # 对分割后的单元格进行汉字检测
                exclude_rows = config.get('exclude_rows', [2])
                _, chinese_cells_info = detect_chinese_in_cells(img_cell_dir, exclude_rows)
                
                # 创建基本文件夹结构
                os.makedirs(img_processed_dir, exist_ok=True)
                
                # 创建包含汉字单元格信息的结果
                results = {}
                if chinese_cells_info:
                    for cell_info in chinese_cells_info:
                        cell_name = cell_info['file_name'].replace('.png', '')
                        results[cell_name] = f"汉字(置信度:{cell_info['confidence']:.2f})"
                else:
                    results = {'chinese_table': f"检测到汉字(表格级置信度:{chinese_detection_results.get('detection_confidence', 0):.2f})"}
                
                skipped_files = {}
                
                # 生成Excel和基础可视化
                img_excel_path = os.path.join(bad_timestamp_dir, f"{img_name}_predictions.xlsx")
                write_results_to_excel(results, img_excel_path, is_combined=False)
                
                img_vis_path = os.path.join(bad_timestamp_dir, f"{img_name}_detection_results.png")
                plt.figure(figsize=(10, 6))
                plt.text(0.5, 0.5, f'检测到汉字表格\n表格级置信度: {chinese_detection_results.get("detection_confidence", 0):.2f}\n单元格级检测: {len(chinese_cells_info)}个汉字单元格', 
                        ha='center', va='center', fontsize=16, transform=plt.gca().transAxes)
                plt.title(f'{img_name} - 汉字检测结果')
                plt.axis('off')
                plt.savefig(img_vis_path, dpi=150, bbox_inches='tight')
                plt.close()
                
                # 保存原始图片
                original_dest = os.path.join(bad_timestamp_dir, f"{img_name}_original.png")
                original_img = cv2.imread(image_path)
                if original_img is not None:
                    cv2.imwrite(original_dest, original_img)
                    print(f"已保存原始图片: {original_dest}")
                
                # 生成汉字单元格拼接图（如果有单元格级检测结果）
                if chinese_cells_info:
                    collage_path = os.path.join(bad_timestamp_dir, f"{img_name}_chinese_cells_collage.png")
                    collage_count = create_chinese_cells_collage(chinese_cells_info, img_cell_dir, collage_path, confidence_threshold=0.6)
                    if collage_count is not None and collage_count > 0:
                        print(f"已生成汉字单元格拼接图: {collage_path} (包含{collage_count}个单元格)")
                else:
                    # 如果没有单元格级检测结果，保存整个表格区域作为汉字检测结果
                    table_collage_path = os.path.join(bad_timestamp_dir, f"{img_name}_chinese_table_region.png")
                    cv2.imwrite(table_collage_path, cropped_cell)
                    print(f"已保存汉字表格区域图: {table_collage_path}")
                
                # 记录表格级汉字过滤原因
                if not hasattr(save_distorted_files, 'filtered_records'):
                    save_distorted_files.filtered_records = []
                
                table_chinese_reason = f"表格级汉字检测(置信度:{chinese_detection_results.get('detection_confidence', 0):.2f})"
                save_distorted_files.filtered_records.append({
                    'img_name': img_name,
                    'filter_reasons': [table_chinese_reason],
                    'filter_time': time.strftime('%Y-%m-%d %H:%M:%S')
                })
                
                chinese_detected_count += 1
                total_processed += 1
                continue

        # 分割单元格
        small_cells = split_into_small_cells(
            cropped_cell,
            output_folder=img_cell_dir
        )

        # 步骤1.5: 基于MNIST的汉字检测（如果启用）
        has_chinese_cells = False
        chinese_cells_info = []
        
        if detection_level == 'cell' and config.get('enable_chinese_detection', False):
            exclude_rows = config.get('exclude_rows', [2])
            print(f"开始MNIST反向汉字检测（排除第{exclude_rows}行）...")
            
            # 先进行图像预处理和数字识别
            processed_count = process_images_for_mnist(
                input_folder=img_cell_dir,
                output_folder=img_processed_dir
            )
            
            if processed_count > 0:
                # 加载模型和transform
                model_file = r"D:\OCR-project\weld_report\model\incremental_nxgs_model.pth"
                if os.path.exists(model_file):
                    model = load_model(model_file)
                    transform = transforms.Compose([
                        transforms.Grayscale(num_output_channels=1),
                        transforms.Resize((28, 28)),
                        transforms.ToTensor(),
                        transforms.Normalize((0.5,), (0.5,))
                    ])
                    
                    # 使用新的MNIST反向检测
                    has_chinese_cells, chinese_cells_info = detect_chinese_by_mnist_reverse(
                        img_processed_dir, model, transform, exclude_rows
                    )
                else:
                    print("模型文件不存在，跳过汉字检测")
            else:
                print("图像预处理失败，跳过汉字检测")
            
            if has_chinese_cells:
                print(f"⚠️ 检测到 {len(chinese_cells_info)} 个单元格包含汉字，分流到待人工处理")
                
                # 创建基本文件夹结构
                os.makedirs(img_processed_dir, exist_ok=True)
                
                # 创建包含汉字单元格信息的结果
                results = {}
                for cell_info in chinese_cells_info:
                    cell_name = cell_info['file_name'].replace('.png', '')
                    results[cell_name] = f"汉字(置信度:{cell_info['confidence']:.2f})"
                
                skipped_files = {}
                
                # 生成Excel和可视化
                img_excel_path = os.path.join(bad_timestamp_dir, f"{img_name}_predictions.xlsx")
                write_results_to_excel(results, img_excel_path, is_combined=False)
                
                img_vis_path = os.path.join(bad_timestamp_dir, f"{img_name}_detection_results.png")
                # 创建可视化，显示包含汉字的单元格
                plt.figure(figsize=(12, 8))
                plt.text(0.5, 0.6, f'检测到汉字单元格\n共{len(chinese_cells_info)}个单元格包含汉字', 
                        ha='center', va='center', fontsize=16, transform=plt.gca().transAxes)
                
                # 显示汉字单元格列表
                cell_list = '\n'.join([f"单元格{info['cell_index']}: 置信度{info['confidence']:.2f}" for info in chinese_cells_info])
                plt.text(0.5, 0.3, cell_list, ha='center', va='center', fontsize=12, 
                        transform=plt.gca().transAxes, bbox=dict(boxstyle="round,pad=0.3", facecolor="lightblue"))
                
                plt.title(f'{img_name} - 单元格汉字检测结果')
                plt.axis('off')
                plt.savefig(img_vis_path, dpi=150, bbox_inches='tight')
                plt.close()
                
                # 保存原始图片
                original_dest = os.path.join(bad_timestamp_dir, f"{img_name}_original.png")
                original_img = cv2.imread(image_path)
                if original_img is not None:
                    cv2.imwrite(original_dest, original_img)
                    print(f"已保存原始图片: {original_dest}")
                
                # 保存单元格网格图到待人工处理文件夹（方便查看哪些单元格有汉字）
                cells_grid_src = os.path.join(img_cell_dir, "cells_grid_with_borders.png")
                if os.path.exists(cells_grid_src):
                    cells_grid_dest = os.path.join(bad_timestamp_dir, f"{img_name}_cells_grid.png")
                    import shutil
                    shutil.copy2(cells_grid_src, cells_grid_dest)
                    print(f"已保存单元格网格图: {cells_grid_dest}")
                
                # 生成汉字单元格拼接图（置信度>0.6）
                collage_path = os.path.join(bad_timestamp_dir, f"{img_name}_chinese_cells_collage.png")
                collage_count = create_chinese_cells_collage(chinese_cells_info, img_cell_dir, collage_path, confidence_threshold=0.6)
                if collage_count is not None and collage_count > 0:
                    print(f"已生成汉字单元格拼接图: {collage_path} (包含{collage_count}个单元格)")
                
                # 记录汉字过滤原因
                if not hasattr(save_distorted_files, 'filtered_records'):
                    save_distorted_files.filtered_records = []
                
                chinese_reason = f"检测到{len(chinese_cells_info)}个汉字单元格"
                save_distorted_files.filtered_records.append({
                    'img_name': img_name,
                    'filter_reasons': [chinese_reason],
                    'filter_time': time.strftime('%Y-%m-%d %H:%M:%S')
                })
                
                print(f"✅ 包含汉字的表格已保存到待人工处理文件夹: {img_name}")
                
                # 更新统计
                chinese_detected_count += 1
                total_processed += 1
                
                # 跳过后续处理
                continue

        # 步骤2: 图像预处理（如果之前没有预处理过）
        if not has_chinese_cells:  # 如果没有进行汉字检测，需要预处理
            processed_count = process_images_for_mnist(
                input_folder=img_cell_dir,
                output_folder=img_processed_dir
            )
        else:  # 如果已经预处理过了，直接获取数量
            processed_count = len([f for f in os.listdir(img_processed_dir) 
                                 if f.startswith('cell_') and f.endswith('.png') and f != 'cell_0.png'])
        
        # 生成预处理步骤对比网格图像
        create_preprocessing_comparison_grids(img_cell_dir, img_processed_dir)

        if processed_count == 0:
            print(f"没有有效图像: {img_name}")
            continue

        # 步骤3: MNIST预测
        model_file = r"D:\OCR-project\weld_report\model\incremental_nxgs_model.pth"

        if not os.path.exists(model_file):
            print("未找到模型文件")
            return

        # 加载模型（如果之前没有加载）
        if 'model' not in locals() or model is None:
            model = load_model(model_file)

        # 定义图像预处理（如果之前没有定义）
        if 'transform' not in locals():
            transform = transforms.Compose([
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((28, 28)),
            transforms.ToTensor(),
            transforms.Normalize((0.5,), (0.5,))
        ])

        # 预测
        results, skipped_files = predict_digits(img_processed_dir, model, transform)
        
        # 显示预测结果统计
        print_detailed_predictions(results, skipped_files)

        if not results and not skipped_files:
            print(f"没有有效结果: {img_name}")
            continue

        # 为每张图片单独生成结果
        img_excel_path = os.path.join(folders['output_dir'], f"{img_name}_predictions.xlsx")
        write_results_to_excel(results, img_excel_path, is_combined=False)

        # 可视化每张图片的结果，包括被跳过的图片
        img_vis_path = os.path.join(folders['output_dir'], f"{img_name}_detection_results.png")
        visualize_results(results, skipped_files, img_processed_dir, img_vis_path)

        # 检查是否有严重扭曲（垂直线和水平线问题）
        has_severe_distortion, distortion_analysis = check_severe_distortion(img_cell_dir)

        # 初始化 filter_reasons
        filter_reasons = distortion_analysis.get('filter_reasons', [])

        if has_severe_distortion:
            # 为扭曲图片生成Excel（保存到Bad文件夹）
            img_excel_path = os.path.join(bad_timestamp_dir, f"{img_name}_predictions.xlsx")
            write_results_to_excel(results, img_excel_path, is_combined=False)
            
            img_vis_path = os.path.join(bad_timestamp_dir, f"{img_name}_detection_results.png")
            visualize_results(results, skipped_files, img_processed_dir, img_vis_path)
            
            # 保存到Bad文件夹
            success = save_distorted_files(
                img_name, image_path, img_excel_path, img_vis_path, 
                img_cell_dir, bad_timestamp_dir, filter_reasons
            )
            
            if success:
                print(f"✅ 已保存到待人工处理文件夹: {os.path.join(bad_timestamp_dir, img_name)}")
        else:
            # 正常图片：保存到待纠偏处理文件夹
            print(f"✅ 正常图片，保存到待纠偏处理文件夹: {img_name}")
            
            # 1. 保存Excel和可视化图到待纠偏处理文件夹
            correction_excel_path = os.path.join(correction_timestamp_dir, f"{img_name}_predictions.xlsx")
            write_results_to_excel(results, correction_excel_path, is_combined=False)
            
            correction_vis_path = os.path.join(correction_timestamp_dir, f"{img_name}_detection_results.png")
            visualize_results(results, skipped_files, img_processed_dir, correction_vis_path)
            
            # 2. 复制原始图片到待纠偏处理文件夹
            correction_original_path = os.path.join(correction_timestamp_dir, f"{img_name}_original.png")
            original_img = cv2.imread(image_path)
            if original_img is not None:
                cv2.imwrite(correction_original_path, original_img)
                print(f"已保存原始图片: {correction_original_path}")

        # 统计不同类型的扭曲（更新统计逻辑）
        has_vertical = "垂直线严重弯曲" in filter_reasons
        has_horizontal = "水平线间距过小" in filter_reasons
        
        # 更新统计计数（汉字检测已经在前面统计过了）
        if has_vertical and has_horizontal:
            both_distorted_count += 1
        elif has_vertical:
            vertical_distorted_count += 1
        elif has_horizontal:
            horizontal_distorted_count += 1
        
        total_processed += 1

        # 简单完成提示
        if results:
            print(f"{img_name}: 识别{len(results)}个数字")
        else:
            print(f"{img_name}: 无识别结果")

        # 清理matplotlib内存 - 每处理完一张图片后清理
        plt.close('all')
        if idx % 10 == 0:
            import gc
            gc.collect()

    # 计算总扭曲数量（包含汉字检测）
    total_distorted = vertical_distorted_count + horizontal_distorted_count + both_distorted_count + chinese_detected_count

    # 如果有被过滤的图片，创建汇总文件和Excel
    if total_distorted > 0:
        # 创建文本汇总文件
        summary_file = os.path.join(bad_timestamp_dir, "待人工处理图片汇总.txt")
        with open(summary_file, 'w', encoding='utf-8') as f:
            f.write(f"待人工处理图片汇总\n")
            f.write(f"过滤时间: {time.strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"总计: {total_distorted}张\n\n")
            f.write("过滤的图片名称:\n")
            
            # 收集被过滤的图片信息
            filtered_images = []
            filtered_count = 0
            
            # 获取过滤记录
            filtered_records = getattr(save_distorted_files, 'filtered_records', [])
            
            for idx, image_path in enumerate(image_paths, 1):
                img_name = os.path.splitext(os.path.basename(image_path))[0]
                
                # 检查该图片是否被过滤
                original_file = os.path.join(bad_timestamp_dir, f"{img_name}_original.png")
                if os.path.exists(original_file):
                    filtered_count += 1
                    
                    # 查找对应的过滤原因
                    filter_reasons = []
                    filter_time = time.strftime('%Y-%m-%d %H:%M:%S')
                    
                    for record in filtered_records:
                        if record['img_name'] == img_name:
                            filter_reasons = record['filter_reasons']
                            filter_time = record['filter_time']
                            break
                    
                    reason_text = ' + '.join(filter_reasons) if filter_reasons else '未知原因'
                    f.write(f"{filtered_count}. {img_name} - {reason_text}\n")
                    
                    filtered_images.append({
                        '序号': filtered_count,
                        '图片名称': img_name,
                        '过滤时间': filter_time,
                        '过滤原因': reason_text
                    })
        
        print(f"过滤图片汇总已保存: {summary_file}")
        
        # 创建Excel汇总文件
        try:
            excel_summary_path = os.path.join(bad_timestamp_dir, "待人工处理图片汇总.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "待人工处理图片汇总"
            
            # 设置表头
            headers = ['序号', '所属PDF', '页码', '过滤时间', '过滤原因']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 写入数据
            for row, img_info in enumerate(filtered_images, 2):
                # 分离图片名称和页数
                full_name = img_info['图片名称']
                if '_' in full_name:
                    parts = full_name.rsplit('_', 1)  # 从右边分割一次
                    img_name = parts[0]
                    page_code = parts[1]
                    # 将页码转换为实际页数（00->1, 01->2, 02->3...）
                    try:
                        page_num = int(page_code) + 1
                    except ValueError:
                        page_num = page_code  # 如果转换失败，保持原值
                else:
                    img_name = full_name
                    page_num = ""
                
                ws.cell(row=row, column=1, value=img_info['序号'])
                ws.cell(row=row, column=2, value=img_name)
                ws.cell(row=row, column=3, value=page_num)
                ws.cell(row=row, column=4, value=img_info['过滤时间'])
                ws.cell(row=row, column=5, value=img_info['过滤原因'])
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8   # 序号
            ws.column_dimensions['B'].width = 30  # 所属PDF
            ws.column_dimensions['C'].width = 8   # 页码
            ws.column_dimensions['D'].width = 20  # 过滤时间
            ws.column_dimensions['E'].width = 40  # 过滤原因
            
            # 设置居中对齐
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            wb.save(excel_summary_path)
            print(f"过滤图片Excel汇总已保存: {excel_summary_path}")
            
        except Exception as e:
            print(f"创建Excel文件时出错: {e}")
            import traceback
            traceback.print_exc()

    # 最终总结
    print(f"\n" + "="*60)
    print(f"处理完成统计:")
    print(f"总处理图片: {total_processed}")
    print(f"问题图片总数: {total_distorted}")
    print(f"  - 仅垂直线扭曲: {vertical_distorted_count}")
    print(f"  - 仅水平线间距问题: {horizontal_distorted_count}")
    print(f"  - 两种问题都有: {both_distorted_count}")
    print(f"  - 包含汉字: {chinese_detected_count}")
    print(f"正常图片: {total_processed - total_distorted}")
    print(f"结果保存至: {folders['output_dir']}")
    
    if total_distorted > 0:
        print(f"问题图片保存至: {bad_timestamp_dir}")
    
    # 如果有汉字检测结果，显示参数调整建议
    if chinese_detected_count > 0:
        print_chinese_detection_tips()
    
    print(f"="*60)

    # 最终清理
    plt.close('all')
    import gc
    gc.collect()


def print_chinese_detection_tips():
    """
    打印手写汉字检测的参数调整建议
    """
    print("\n" + "="*60)
    print("手写汉字检测参数调整建议:")
    print("="*60)
    
    detection_level = CHINESE_DETECTION_CONFIG.get('detection_level', 'table')
    print(f"当前检测级别: {detection_level}")
    print(f"  - table: 在整个表格级别检测汉字")
    print(f"  - cell: 在单个单元格级别检测汉字（推荐）")
    print()
    
    if detection_level == 'cell':
        exclude_rows = CHINESE_DETECTION_CONFIG.get('exclude_rows', [2])
        confidence_threshold = CHINESE_DETECTION_CONFIG.get('confidence_threshold', 0.75)
        min_fg = CHINESE_DETECTION_CONFIG.get('min_foreground_density', 0.15)
        max_fg = CHINESE_DETECTION_CONFIG.get('max_foreground_density', 0.75)
        min_area = CHINESE_DETECTION_CONFIG.get('min_area_threshold', 80)
        
        print("单元格级别检测特点（已优化）:")
        print("  ✓ 智能数字过滤：专门区分数字和汉字，减少误检")
        print("  ✓ 多特征分析：连通域数量、前景密度、宽高比、复杂度等")
        print("  ✓ 数字惩罚机制：对数字特征进行负向评分")
        print("  ✓ 局部变化分析：检测笔画密度分布的复杂性")
        print(f"  ✓ 智能过滤：自动排除第{exclude_rows}行（通常是表头）")
        print()
        print("当前平衡参数:")
        print(f"  - 置信度阈值: {confidence_threshold*100:.0f}%（平衡点，可上调过滤数字）")
        print(f"  - 前景密度范围: {min_fg:.2f} - {max_fg:.2f}（放宽以包含更多汉字）")
        print(f"  - 最小面积: {min_area}像素（降低以包含小汉字）")
        print()
        print("新策略特性:")
        print("  ✓ 宽容汉字检测：降低各项要求，减少汉字漏检")
        print("  ✓ 置信度区分：数字误检但置信度低，真汉字置信度高")
        print("  ✓ 中心区域聚焦：避免边缘表格线干扰")
        print("  ✓ 9维数字过滤 + 8维汉字特征：保持精确分析")
        print()
        print("调整策略:")
        print("🎯 推荐做法：观察拼接图中的置信度分布")
        print("  - 真汉字通常置信度 > 0.70")
        print("  - 误检数字通常置信度 < 0.70")
        print()
        print("🔧 如果数字/字母仍被误识别（但置信度应该不高）:")
        print(f"  1. 提高置信度阈值: confidence_threshold = 0.75 (当前{confidence_threshold})")
        print(f"  2. 或更高: confidence_threshold = 0.80")
        print("  3. 开启调试信息: enable_debug_info = True")
        print()
        print("🔧 如果汉字仍然漏检:")
        print(f"  1. 降低置信度阈值: confidence_threshold = 0.60 (当前{confidence_threshold})")
        print(f"  2. 降低前景密度下限: min_foreground_density = 0.10 (当前{min_fg})")
        print(f"  3. 降低最小面积: min_area_threshold = 40 (当前{min_area})")
        print("  4. 开启调试查看被漏检的汉字: enable_debug_info = True")
        print()
        print("🔧 排除行设置:")
        print(f"  - 当前排除: 第{exclude_rows}行")
        print("  - 排除更多行: exclude_rows = [1, 2] （第1和第2行）")
        print("  - 不排除任何行: exclude_rows = []")
    else:
        print("表格级别检测参数:")
        print("如果检测过于敏感（误识别数字为汉字）:")
        print("  - 增加 min_char_components (当前: {})".format(CHINESE_DETECTION_CONFIG['min_char_components']))
        print("  - 增加 min_foreground_ratio (当前: {})".format(CHINESE_DETECTION_CONFIG['min_foreground_ratio']))
        print("  - 增加 min_edge_density (当前: {})".format(CHINESE_DETECTION_CONFIG['min_edge_density']))
        print("  - 增加 min_local_complexity (当前: {})".format(CHINESE_DETECTION_CONFIG['min_local_complexity']))
        print("  - 增加 min_indicators (当前: {})".format(CHINESE_DETECTION_CONFIG['min_indicators']))
        print()
        print("如果检测不够敏感（漏检手写汉字）:")
        print("  - 减少上述参数值")
    
    print()
    print("检测级别切换:")
    print("  - 改为单元格级别：设置 detection_level = 'cell'（推荐）")
    print("  - 改为表格级别：设置 detection_level = 'table'")
    print("  - 完全关闭检测：设置 enable_chinese_detection = False")
    print("="*60)


if __name__ == "__main__":

    # 设置命令行参数
    parser = argparse.ArgumentParser(description='OCR处理程序 - 使用YOLO检测表格、预处理图像并预测手写数字和字母(0-9+N,X,G,S)，支持汉字检测分流')
    parser.add_argument('input_path', type=str, help='输入图像的路径或包含图像的目录')
    
    args = parser.parse_args()
    
    # YOLO模型路径固定在代码中
    yolo_model_path = r'D:\OCR-project\weld_report\best.pt'  # YOLO模型路径

    # 检查输入路径
    if not os.path.exists(args.input_path):
        print(f"错误: 输入路径不存在")
        sys.exit(1)

    # 检查YOLO模型文件
    if not os.path.exists(yolo_model_path):
        print(f"错误: YOLO模型文件不存在")
        sys.exit(1)
    main(args.input_path, yolo_model_path)

