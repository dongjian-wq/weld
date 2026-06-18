import os
import sys
import time
import cv2
import torch
import numpy as np
import torchvision.transforms as transforms
from torch import nn
from PIL import Image
import openpyxl
from openpyxl.styles import Alignment
import matplotlib.pyplot as plt
from datetime import datetime

# 配置matplotlib支持中文显示
plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'SimSun', 'Arial Unicode MS',
                                   'DejaVu Sans']  # 优先使用的字体系列
plt.rcParams['axes.unicode_minus'] = False  # 解决保存图像时负号'-'显示为方块的问题
plt.rcParams['font.family'] = 'sans-serif'  # 使用sans-serif字体


# 设置中文字体的函数
def set_chinese_font():
    """
    尝试设置中文字体，针对不同操作系统选择不同字体
    """
    import platform
    system = platform.system()

    if system == 'Windows':
        fonts = ['SimHei', 'Microsoft YaHei', 'SimSun']
    elif system == 'Darwin':  # macOS
        fonts = ['PingFang SC', 'Heiti SC', 'STHeiti']
    else:  # Linux等其他系统
        fonts = ['WenQuanYi Zen Hei', 'WenQuanYi Micro Hei', 'Droid Sans Fallback']

    # 尝试设置字体
    for font in fonts:
        try:
            plt.rcParams['font.sans-serif'] = [font] + plt.rcParams['font.sans-serif']
            # 测试字体是否可用
            fig = plt.figure(figsize=(1, 1))
            plt.text(0.5, 0.5, '测试', fontsize=12, ha='center', va='center')
            plt.close(fig)
            print(f"成功设置中文字体: {font}")
            return True
        except Exception as e:
            continue

    print("警告: 未能找到合适的中文字体，图像中的中文可能不会正确显示")
    return False


# 在程序开始时调用此函数
set_chinese_font()


def create_timestamped_folders():
    """
    创建带时间戳的文件夹路径
    """
    base_dir = os.path.dirname(os.path.abspath(__file__))
    timestamp = datetime.now().strftime("%Y-%m-%d_%H.%M.%S")

    # 创建各个处理阶段的输出文件夹
    data_dir = os.path.join(base_dir, "data")
    cell_dir = os.path.join(data_dir, "cell", timestamp)
    processed_cells_dir = os.path.join(data_dir, "processed_cells", timestamp)
    output_dir = os.path.join(base_dir, "处理结果", timestamp)

    # 确保各个文件夹存在
    for dir_path in [data_dir, os.path.dirname(cell_dir), os.path.dirname(processed_cells_dir),
                     os.path.dirname(output_dir), cell_dir, processed_cells_dir, output_dir]:
        os.makedirs(dir_path, exist_ok=True)

    return {
        "timestamp": timestamp,
        "cell_dir": cell_dir,
        "processed_cells_dir": processed_cells_dir,
        "output_dir": output_dir
    }


def detect_and_crop_largest_table_cell(image_path, output_folder, min_cell_size=10, crop_left=0.04, crop_right=0.5,
                                       crop_top=0.2, crop_bottom=0.15):
    """
    检测表格线并切割最大的表格单元格区域
    """
    print(f"处理图像: {image_path}")
    original = cv2.imread(image_path)
    if original is None:
        raise FileNotFoundError(f"无法读取图像: {image_path}")

    gray = cv2.cvtColor(original, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, -2)

    horizontal = thresh.copy()
    vertical = thresh.copy()
    rows, cols = horizontal.shape
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (cols // 50, 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, rows // 50))
    horizontal = cv2.dilate(cv2.erode(horizontal, horizontal_kernel), horizontal_kernel)
    vertical = cv2.dilate(cv2.erode(vertical, vertical_kernel), vertical_kernel)

    table_lines = cv2.bitwise_or(horizontal, vertical)

    contours, _ = cv2.findContours(table_lines, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    # 初始化最大面积和对应的轮廓
    max_area = 0
    largest_cell = None

    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)
        area = w * h

        # 跳过小于最小尺寸的单元格
        if w < min_cell_size or h < min_cell_size:
            continue

        # 更新最大面积和对应的单元格
        if area > max_area:
            max_area = area
            largest_cell = (x, y, w, h)

    if largest_cell is not None:
        x, y, w, h = largest_cell
        first_cell = original[y:y + h, x:x + w]

        # 保存裁剪前的最大单元格图像到输出文件夹
        cv2.imwrite(os.path.join(output_folder, "largest_cell.png"), first_cell)

        new_x = int(w * crop_left)
        new_y = int(h * crop_top)
        new_w = int(w * (1 - crop_left - crop_right))
        new_h = int(h * (1 - crop_top - crop_bottom))
        new_w = max(1, new_w)
        new_h = max(1, new_h)

        cropped_cell = first_cell[new_y:new_y + new_h, new_x:new_x + new_w]
        return cropped_cell

    return None


def split_into_small_cells(cropped_cell, output_folder, min_cell_size=10, max_cell_height=90):
    """
    将大的表格区域分割为小单元格
    """
    gray = cv2.cvtColor(cropped_cell, cv2.COLOR_BGR2GRAY)
    thresh = cv2.adaptiveThreshold(~gray, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 15, -2)

    horizontal = thresh.copy()
    vertical = thresh.copy()
    rows, cols = horizontal.shape
    horizontal_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (cols // 20, 1))
    vertical_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, rows // 20))
    horizontal = cv2.dilate(cv2.erode(horizontal, horizontal_kernel), horizontal_kernel)
    vertical = cv2.dilate(cv2.erode(vertical, vertical_kernel), vertical_kernel)

    table_lines = cv2.bitwise_or(horizontal, vertical)

    contours, _ = cv2.findContours(table_lines, cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)

    small_cells = []
    cell_coords = []  # 用于存储每个单元格的坐标

    for cnt in contours:
        x, y, w, h = cv2.boundingRect(cnt)

        if w < min_cell_size or h < min_cell_size:
            continue

        if h <= max_cell_height:
            cell_img = cropped_cell[y:y + h, x:x + w]

            # 当宽度/高度 > 1.5 时进行拆分
            if w / h > 1.5:
                num_splits = round(w / h, 1)  # 先保留一位小数
                num_splits = round(num_splits)  # 再四舍五入

                num_splits = max(1, num_splits)  # 确保至少拆分 1 份
                split_width = w // num_splits  # 计算每个子单元格宽度

                for i in range(num_splits):
                    split_x = x + i * split_width
                    split_cell = cropped_cell[y:y + h, split_x:split_x + split_width]
                    small_cells.append(split_cell)
                    cell_coords.append((split_x, y, split_width, h))  # 记录坐标
            else:
                small_cells.append(cell_img)
                cell_coords.append((x, y, w, h))  # 记录坐标

    # 对单元格按照 (y, x) 排序
    tolerance = 15  # 容差值，单位为像素
    sorted_cells = sorted(
        zip(cell_coords, small_cells),
        key=lambda item: (round(item[0][1] / tolerance) * tolerance, item[0][0])  # 按 Y 坐标分组，再按 X 坐标排序
    )

    cell_coords, small_cells = zip(*sorted_cells) if sorted_cells else ([], [])

    # 保存所有单元格
    for i, cell in enumerate(small_cells):
        out_path = os.path.join(output_folder, f"cell_{i}.png")
        cv2.imwrite(out_path, cell)

    print(f"共检测到 {len(small_cells)} 个单元格，保存到 {output_folder}")

    return list(small_cells)


def crop_border(image, border_ratio=0.1):
    """
    裁剪图片边缘
    """
    h, w = image.shape[:2] if len(image.shape) > 2 else image.shape
    crop_h, crop_w = int(h * border_ratio), int(w * border_ratio)

    if len(image.shape) > 2:
        cropped = image[crop_h:h - crop_h, crop_w:w - crop_w]
    else:
        cropped = image[crop_h:h - crop_h, crop_w:w - crop_w]

    return cropped


def preprocess_image(image, target_size=(28, 28), padding=4):
    """
    对图像进行预处理：裁剪边框、二值化、调整大小
    """
    # 转为灰度图
    if len(image.shape) > 2:
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    else:
        gray = image

    # 裁剪边框
    cropped = crop_border(gray, border_ratio=0.09)

    # 转为PIL图像进行进一步处理
    pil_img = Image.fromarray(cropped)

    # 转为灰度图
    pil_img = pil_img.convert('L')

    # 二值化（黑白图）
    pil_img = pil_img.point(lambda x: 0 if x < 128 else 255, '1')

    # 反转颜色（黑底白字）
    pil_img = ImageOps.invert(pil_img.convert('L'))

    # 裁剪字符边缘
    bbox = pil_img.getbbox()
    if bbox:
        pil_img = pil_img.crop(bbox)

    # 添加一些边距避免贴边
    pil_img = ImageOps.expand(pil_img, border=padding, fill=0)

    # 等比例缩放
    pil_img.thumbnail((target_size[0], target_size[1]), Image.Resampling.LANCZOS)

    # 创建黑底图像并将缩放后的图片粘贴居中
    new_img = Image.new('L', target_size, 0)
    upper_left = ((target_size[0] - pil_img.size[0]) // 2, (target_size[1] - pil_img.size[1]) // 2)
    new_img.paste(pil_img, upper_left)

    return new_img


def process_images_for_mnist(input_folder, output_folder):
    """
    处理所有单元格图像，使其适合MNIST模型识别
    """
    os.makedirs(output_folder, exist_ok=True)
    print(f"处理图像从 {input_folder} 到 {output_folder}")

    # 获取并排序输入文件
    file_list = []
    for file_name in os.listdir(input_folder):
        # 过滤掉非图片文件
        if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            continue

        # 过滤掉中间处理文件
        if file_name in ['cropped_cell.png', 'largest_cell.png']:
            # 仍然复制这些文件，但不进行处理
            input_path = os.path.join(input_folder, file_name)
            output_path = os.path.join(output_folder, file_name)
            try:
                image = cv2.imread(input_path)
                if image is not None:
                    cv2.imwrite(output_path, image)
                    print(f"已复制中间处理文件: {file_name}")
            except Exception as e:
                print(f"复制 {file_name} 时出错: {e}")
            continue

        # 只处理cell_开头的文件
        if not file_name.startswith('cell_'):
            print(f"跳过非单元格文件: {file_name}")
            continue

        file_list.append(file_name)

    # 对文件进行数值排序
    def get_file_index(filename):
        # 针对"cell_数字.png"格式的文件提取数字部分
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')  # 对于无法解析的文件名放在最后

    file_list.sort(key=get_file_index)

    processed_count = 0
    for file_name in file_list:
        input_path = os.path.join(input_folder, file_name)
        output_path = os.path.join(output_folder, file_name)

        try:
            # 读取图像
            image = cv2.imread(input_path)
            if image is None:
                print(f"无法读取图片: {input_path}")
                continue

            # 处理图像
            processed_img = preprocess_image(image)
            processed_img.save(output_path)
            processed_count += 1

        except Exception as e:
            print(f"处理 {file_name} 时出错: {e}")

    print(f"成功处理 {processed_count} 张图像")
    return processed_count


def load_model(model_path):
    """
    加载预训练的 MNIST 模型
    """
    class CNN(nn.Module):
        def __init__(self):
            super(CNN, self).__init__()
            # 第一个卷积块
            self.conv1 = nn.Sequential(
                nn.Conv2d(1, 32, kernel_size=3, padding=1),
                nn.BatchNorm2d(32),
                nn.ReLU(),
                nn.Conv2d(32, 64, kernel_size=3, padding=1),
                nn.BatchNorm2d(64),
                nn.ReLU(),
                nn.MaxPool2d(2)
            )
            
            # 第二个卷积块
            self.conv2 = nn.Sequential(
                nn.Conv2d(64, 128, kernel_size=3, padding=1),
                nn.BatchNorm2d(128),
                nn.ReLU(),
                nn.Conv2d(128, 128, kernel_size=3, padding=1),
                nn.BatchNorm2d(128),
                nn.ReLU(),
                nn.MaxPool2d(2)
            )
            
            # 全连接层
            self.fc = nn.Sequential(
                nn.Flatten(),
                nn.Linear(128 * 7 * 7, 256),
                nn.ReLU(),
                nn.Dropout(0.4),
                nn.Linear(256, 128),
                nn.ReLU(),
                nn.Dropout(0.3),
                nn.Linear(128, 15)
            )

        def forward(self, x):
            x = self.conv1(x)
            x = self.conv2(x)
            x = self.fc(x)
            return x

    model = CNN()
    model.load_state_dict(torch.load(model_path, map_location=torch.device('cuda')))
    model.eval()
    return model


def predict_digits(folder_path, model, transform):
    """
    使用模型预测文件夹中的手写数字和字母，筛选掉接近纯黑的图片
    """
    # 定义标签映射字典
    label_map = {
        0: '0', 1: '1', 2: '2', 3: '3', 4: '4',
        5: '5', 6: '6', 7: '7', 8: '8', 9: '9',
        10: 'C', 11: 'D', 12: 'S', 13: 'Z', 14: 'X'  # 将X和Z的位置对调
    }
    
    results = {}
    skipped_files = {}  # 存储被跳过的文件及原因

    # 获取文件列表
    file_list = []
    for file_name in os.listdir(folder_path):
        # 过滤掉非图片文件
        if not file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
            continue

        # 过滤掉中间处理文件
        if file_name in ['cropped_cell.png', 'largest_cell.png']:
            print(f"跳过中间处理文件: {file_name}")
            continue

        # 只处理cell_开头的文件
        if not file_name.startswith('cell_'):
            print(f"跳过非单元格文件: {file_name}")
            continue

        file_list.append(file_name)

    # 对文件进行数值排序
    def get_file_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')

    # 按数值排序文件列表
    file_list.sort(key=get_file_index)

    # 处理排序后的文件
    for file_name in file_list:
        file_path = os.path.join(folder_path, file_name)

        # 读取图片并转换为灰度图
        image = cv2.imread(file_path, cv2.IMREAD_GRAYSCALE)
        if image is None:
            print(f"无法读取图片: {file_path}")
            skipped_files[file_name] = "无法读取图片"
            continue

        # 检查图片是否接近纯黑
        if np.mean(image) < 5 or np.count_nonzero(image > 10) < image.shape[0] * image.shape[1] * 0.01:
            print(f"图片 {file_name} 被检测为接近纯黑，跳过预测")
            skipped_files[file_name] = "接近纯黑"
            continue

        # 转换为 PIL 格式并预处理
        image = Image.fromarray(image)
        image = transform(image).unsqueeze(0)  # 添加 batch 维度

        # 预测
        with torch.no_grad():
            output = model(image)
            predicted_idx = torch.argmax(output, dim=1).item()
            predicted_label = label_map[predicted_idx]
            results[file_name] = predicted_label
            print(f"图片 {file_name} 预测结果: {predicted_label} (索引: {predicted_idx})")

    return results, skipped_files


def write_results_to_excel(results, excel_path, is_combined=False):
    """
    将预测结果写入 Excel 文件，按照固定的 10 行 16 列顺序排列
    :param results: 预测结果字典 {文件名: 预测值}
    :param excel_path: 输出 Excel 文件路径
    :param is_combined: 是否为合并结果表格
    """
    # 创建 Excel 工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "预测结果"

    # 写入预测结果
    for file_name, digit in results.items():
        try:
            # 获取文件索引
            if is_combined:
                # 合并结果文件名格式为 "图片名_cell_数字.png"
                parts = file_name.split('_')
                if len(parts) >= 3 and parts[-1].endswith('.png'):
                    img_name = parts[0]
                    # 提取 "cell_数字" 中的数字部分
                    index_str = parts[-1].split('.')[0]  # 去掉.png
                    if index_str.isdigit():
                        index = int(index_str)
                    else:
                        print(f"警告: 无法从 {file_name} 中提取有效索引，跳过")
                        continue
                else:
                    print(f"警告: 文件名 {file_name} 格式不正确，跳过")
                    continue
            else:
                # 单图片结果，文件名格式为 "cell_数字.png"
                if not file_name.startswith("cell_"):
                    print(f"警告: 文件名 {file_name} 格式不正确，跳过")
                    continue
                index = int(file_name.split("_")[1].split(".")[0])  # 从文件名中提取单元格索引
                img_name = ""  # 单图片处理时不需要标记图片名

            # 计算行列位置
            row = (index // 16) + 1  # 每 16 个单元格换一行
            col = (index % 16) + 1  # 每行最多 16 列

            # 写入 Excel 单元格
            if is_combined and img_name:
                ws.cell(row=row, column=col, value=f"{img_name}_{index}: {digit}")
            else:
                ws.cell(row=row, column=col, value=f"{file_name}: {digit}")

        except (IndexError, ValueError) as e:
            print(f"警告: 处理文件名 {file_name} 时出错: {e}")
            continue

    # 设置列宽以确保内容完全显示
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # 获取列字母
        for cell in col:
            if cell.value:
                # 计算单元格内容的显示长度（考虑中文字符）
                try:
                    cell_length = 0
                    for char in str(cell.value):
                        # 中文字符计为2个单位宽度，其他字符计为1个单位宽度
                        if ord(char) > 127:
                            cell_length += 2
                        else:
                            cell_length += 1
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    # 发生错误时使用字符串长度
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))

        # 设置列宽，增加一些额外空间
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # 设置单元格居中
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 保存 Excel 文件
    wb.save(excel_path)
    print(f"预测结果已保存到 {excel_path}")


def visualize_results(results, skipped_files, input_folder, output_path):
    """
    可视化预测结果，包括成功预测和被跳过的图片，使用固定的16列布局
    :param results: 预测结果字典 {文件名: 预测值}
    :param skipped_files: 被跳过的文件字典 {文件名: 跳过原因}
    :param input_folder: 输入图片文件夹路径
    :param output_path: 输出可视化图片路径
    """
    # 合并所有文件，包括成功预测和跳过的
    all_files = list(results.keys()) + list(skipped_files.keys())

    # 定义文件排序函数
    def get_file_index(filename):
        try:
            return int(filename.split("_")[1].split(".")[0])
        except (ValueError, IndexError):
            return float('inf')

    # 按数字顺序排序所有文件
    all_files.sort(key=get_file_index)

    n_samples = len(all_files)
    if n_samples == 0:
        print("没有图片可视化")
        return

    # 固定为16列布局
    fixed_cols = 16
    rows = (n_samples + fixed_cols - 1) // fixed_cols

    # 创建画布，设置更大的尺寸以适应16列
    fig_width = fixed_cols * 1.2  # 每列1.2英寸宽
    fig_height = rows * 1.2  # 每行1.2英寸高
    fig, axes = plt.subplots(rows, fixed_cols, figsize=(fig_width, fig_height))

    # 确保axes始终是二维数组
    if rows == 1:
        axes = axes.reshape(1, -1)

    # 绘制每个图片，按行列位置排列
    for i, file_name in enumerate(all_files):
        # 将单元格索引转换为行列位置
        file_index = get_file_index(file_name)
        if file_index == float('inf'):  # 对于无法解析索引的文件
            row_idx = i // fixed_cols
            col_idx = i % fixed_cols
        else:
            # 让单元格按原始表格位置排列
            row_idx = file_index // fixed_cols
            col_idx = file_index % fixed_cols

        # 确保不超出画布范围
        if row_idx >= rows or col_idx >= fixed_cols:
            print(f"警告: 文件 {file_name} (索引: {file_index}) 超出画布范围")
            continue

        # 读取图像
        image_path = os.path.join(input_folder, file_name)
        image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)

        if image is None:
            # 处理无法读取的图片
            axes[row_idx, col_idx].text(0.5, 0.5, "图片加载失败",
                                        ha='center', va='center', fontsize=7)
            axes[row_idx, col_idx].set_title(f"#{file_index}", fontsize=7)
        else:
            # 显示图像
            axes[row_idx, col_idx].imshow(image, cmap='gray')

            # 设置标题
            if file_name in results:
                axes[row_idx, col_idx].set_title(f"#{file_index}: {results[file_name]}",
                                                 fontsize=7, color='blue')
            else:
                reason = skipped_files.get(file_name, "未知原因")
                # 使用简短的原因说明，以适应小尺寸标题
                short_reason = "黑" if reason == "接近纯黑" else "错误"
                axes[row_idx, col_idx].set_title(f"#{file_index}: {short_reason}",
                                                 fontsize=7, color='red')

        axes[row_idx, col_idx].axis('off')

    # 隐藏未使用的子图
    for r in range(rows):
        for c in range(fixed_cols):
            # 检查此位置是否有对应文件显示
            has_image = False
            for file_name in all_files:
                file_index = get_file_index(file_name)
                if file_index != float('inf'):
                    if file_index // fixed_cols == r and file_index % fixed_cols == c:
                        has_image = True
                        break

            # 若无图像显示，则隐藏坐标轴
            if not has_image:
                axes[r, c].axis('off')

    plt.tight_layout()
    plt.subplots_adjust(hspace=0.4)  # 增加行间距，避免标题重叠
    plt.savefig(output_path, dpi=150)  # 使用更高的DPI获得更清晰的图像
    print(f"可视化结果(16列布局)已保存到 {output_path}")


def main(input_path):
    """
    主函数，整合各个处理步骤
    """
    print("=" * 50)
    print("开始处理OCR任务")
    print("=" * 50)

    start_time = time.time()

    # 1. 创建带时间戳的文件夹
    folders = create_timestamped_folders()
    print(f"使用时间戳: {folders['timestamp']}")
    print(f"表格分割输出目录: {folders['cell_dir']}")
    print(f"图像预处理输出目录: {folders['processed_cells_dir']}")
    print(f"最终结果输出目录: {folders['output_dir']}")

    # 处理图片路径
    image_paths = []
    if os.path.isdir(input_path):
        # 如果输入是目录，获取所有图片
        for file_name in os.listdir(input_path):
            if file_name.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp')):
                image_paths.append(os.path.join(input_path, file_name))
        print(f"在目录 {input_path} 中找到 {len(image_paths)} 张图片")
    else:
        # 如果输入是单个文件
        image_paths = [input_path]

    if not image_paths:
        print(f"在指定路径 {input_path} 未找到任何图片")
        return

    # 2. 处理每张图片
    for idx, image_path in enumerate(image_paths):
        print(f"\n处理图片 {idx + 1}/{len(image_paths)}: {os.path.basename(image_path)}")

        # 为每张图片创建子文件夹
        img_name = os.path.splitext(os.path.basename(image_path))[0]
        img_cell_dir = os.path.join(folders['cell_dir'], img_name)
        img_processed_dir = os.path.join(folders['processed_cells_dir'], img_name)
        os.makedirs(img_cell_dir, exist_ok=True)
        os.makedirs(img_processed_dir, exist_ok=True)

        # 步骤1: 检测表格线并分割单元格
        print(f"\n[步骤1] 检测表格线并分割单元格: {img_name}")
        cropped_cell = detect_and_crop_largest_table_cell(
            image_path,
            output_folder=img_cell_dir,
            min_cell_size=20,
            crop_left=0.02,
            crop_right=0.524,
            crop_top=0.324,
            crop_bottom=0.18
        )

        if cropped_cell is None:
            print(f"未检测到表格: {img_name}")
            continue

        # 保存裁剪后的表格
        cv2.imwrite(os.path.join(img_cell_dir, "cropped_cell.png"), cropped_cell)

        # 分割单元格
        small_cells = split_into_small_cells(
            cropped_cell,
            output_folder=img_cell_dir,
            min_cell_size=20,
            max_cell_height=90
        )

        step1_time = time.time()

        # 步骤2: 图像预处理
        print(f"\n[步骤2] 图像预处理: {img_name}")
        processed_count = process_images_for_mnist(
            input_folder=img_cell_dir,
            output_folder=img_processed_dir
        )

        if processed_count == 0:
            print(f"没有有效图像进行预处理: {img_name}")
            continue

        step2_time = time.time()

        # 步骤3: MNIST预测
        print(f"\n[步骤3] MNIST预测手写数字: {img_name}")

        # 使用具体路径
        model_file = r"D:\OCR-project\weld_report\model\custom_mnist_15class.pth"
        print(f"使用模型: {model_file}")

        if not os.path.exists(model_file):
            print("未找到模型文件，无法进行预测")
            return

        # 加载模型
        model = load_model(model_file)

        # 定义图像预处理
        transform = transforms.Compose([
            transforms.Grayscale(num_output_channels=1),
            transforms.Resize((28, 28)),
            transforms.ToTensor(),
            transforms.Normalize((0.5,), (0.5,))
        ])

        # 预测
        results, skipped_files = predict_digits(img_processed_dir, model, transform)

        if not results and not skipped_files:
            print(f"没有有效的预测结果或跳过文件: {img_name}")
            continue

        # 为每张图片单独生成结果
        img_excel_path = os.path.join(folders['output_dir'], f"{img_name}_predictions.xlsx")
        write_results_to_excel(results, img_excel_path, is_combined=False)

        # 可视化每张图片的结果，包括被跳过的图片
        img_vis_path = os.path.join(folders['output_dir'], f"{img_name}_detection_results.png")
        visualize_results(results, skipped_files, img_processed_dir, img_vis_path)

        # 复制中间处理文件到处理结果目录
        for intermediate_file in ['cropped_cell.png', 'largest_cell.png']:
            source_path = os.path.join(img_cell_dir, intermediate_file)
            if os.path.exists(source_path):
                # 添加图片名称标识
                dest_filename = f"{img_name}_{intermediate_file}"
                dest_path = os.path.join(folders['output_dir'], dest_filename)
                try:
                    # 读取图像并添加图片名称标识
                    img = cv2.imread(source_path)
                    if img is not None:
                        # 添加文字标识
                        font = cv2.FONT_HERSHEY_SIMPLEX
                        text = f"图片: {img_name}"
                        h, w = img.shape[:2]

                        # 计算合适的字体大小和位置
                        font_scale = min(w, h) * 0.001  # 根据图像大小调整字体大小
                        font_scale = max(0.5, min(font_scale, 2.0))  # 限制字体大小范围
                        thickness = max(1, int(font_scale * 2))

                        # 在图像顶部添加背景条
                        cv2.rectangle(img, (0, 0), (w, int(h * 0.06)), (255, 255, 255), -1)

                        # 添加文字
                        cv2.putText(img, text, (10, int(h * 0.04)), font, font_scale, (0, 0, 0), thickness)

                        # 保存添加了标识的图像
                        cv2.imwrite(dest_path, img)
                        print(f"已复制并标记中间处理文件: {dest_filename}")
                    else:
                        print(f"无法读取中间处理文件: {source_path}")
                except Exception as e:
                    print(f"处理中间文件 {intermediate_file} 时出错: {e}")

        print(f"图片 {img_name} 处理完成, 识别到 {len(results)} 个数字")

    # 打印总结
    if len(image_paths) > 1:
        print("\n" + "=" * 50)
        print(f"所有图片处理完成！总共处理 {len(image_paths)} 张图片")
        print(f"每张图片的处理结果已单独保存至: {folders['output_dir']}")
        print("=" * 50)
    elif len(image_paths) == 1:
        img_name = os.path.splitext(os.path.basename(image_paths[0]))[0]
        print("\n" + "=" * 50)
        print(f"图片 {img_name} 处理完成")
        print(f"结果已保存至: {folders['output_dir']}")
        print("=" * 50)


if __name__ == "__main__":
    import argparse
    from PIL import ImageOps

    parser = argparse.ArgumentParser(description='OCR处理程序 - 检测表格线、预处理图像并预测手写数字')
    parser.add_argument('input_path', type=str, help='输入图像的路径或包含图像的目录')

    args = parser.parse_args()

    if not os.path.exists(args.input_path):
        print(f"错误: 输入路径 {args.input_path} 不存在")
        sys.exit(1)

    main(args.input_path)